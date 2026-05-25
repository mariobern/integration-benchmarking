#!/usr/bin/env python3
"""Apply a dq_summary 'allowed' sheet to after.json.

Reads the 'allowed' sheet of a dq_summary_<cluster>_<date>.xlsx (produced by
lazer_dq/summarize_feeds.py) and edits a Lazer config (after.json / after_1.json)
in place: per-(feed, session) it promotes COMING_SOON feeds to STABLE on their
DQ-vetted publisher lists, and additively adds missing sessions to already-live
(STABLE) feeds without disturbing their live sessions.

Run:
    python3 -m lazer_dq.apply_allowed_to_config \
        --xlsx dq_summary_lazer-prod_2026-05-20.xlsx \
        --config after_1.json --dry-run

See docs/apply_allowed_to_config.md and
docs/superpowers/specs/2026-05-26-apply-dq-summary-to-config-design.md.
"""
import argparse
import json
import re
import shutil
import sys
from pathlib import Path

from lib.json_surgery import find_feed_block, find_session_block  # noqa: F401

# Session names, in after.json order.
SESSION_ORDER = ["REGULAR", "PRE_MARKET", "POST_MARKET", "OVER_NIGHT"]

# Publisher 0 (aggregate sentinel) + Lazer publishers. summarize_feeds excludes
# {0} ∪ .Test but NOT Lazer ids, so we strip them defensively here.
EXCLUDED_PUBLISHERS = {0, 1, 9, 13, 15}

# minPublishers defaults per session.
SESSION_MIN_PUBLISHERS = {
    "REGULAR": 3,
    "PRE_MARKET": 2,
    "POST_MARKET": 2,
    "OVER_NIGHT": 1,
}
# REGULAR sessions with this many or fewer publishers use a reduced floor.
REGULAR_LOW_PUB_THRESHOLD = 5
REGULAR_LOW_PUB_MIN = 2


def filter_publishers(ids: list[int]) -> tuple[list[int], list[int]]:
    """Drop EXCLUDED_PUBLISHERS. Return (kept_sorted_unique, removed_sorted)."""
    id_set = set(ids)
    kept = sorted(id_set - EXCLUDED_PUBLISHERS)
    removed = sorted(id_set & EXCLUDED_PUBLISHERS)
    return kept, removed


def get_min_publishers(session: str, pub_count: int) -> int:
    """minPublishers for a session, applying the REGULAR low-count reduction."""
    if session == "REGULAR" and pub_count <= REGULAR_LOW_PUB_THRESHOLD:
        return REGULAR_LOW_PUB_MIN
    return SESSION_MIN_PUBLISHERS[session]


def _parse_ids_cell(cell) -> list[int] | None:
    """Extract publisher ids from an 'allowedPublisherIds' cell.

    The cell is either '(no data)'/None or the paste-ready fragment
    '"allowedPublisherIds": [ 41, 69 ],'. Returns a list of ints, or None
    when there is no list.
    """
    if cell is None:
        return None
    text = str(cell)
    if not text.startswith('"allowedPublisherIds"'):
        return None
    m = re.search(r"\[(.*?)\]", text)
    if not m:
        return None
    inner = m.group(1).strip()
    if not inner:
        return []
    return [int(x) for x in inner.split(",") if x.strip()]


def parse_allowed_sheet(path) -> dict[int, dict]:
    """Parse the 'allowed' sheet, grouped by feed_id.

    Returns {feed_id: {"aggregate": list[int]|None,
                       "sessions": {SESSION: list[int]|None}}}.
    Rows that are not genuine data rows are skipped: those whose Feed ID column
    is not an int (title, header, dividers), and the bare-integer "Feeds skipped"
    footer rows (whose Session cell is empty).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "allowed" not in wb.sheetnames:
        raise ValueError(f"workbook {path} has no 'allowed' sheet")
    ws = wb["allowed"]

    feeds: dict[int, dict] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not isinstance(row[0], int):
            continue
        session = row[1]
        # Only genuine data rows register a feed. This ignores the bare-integer
        # "Feeds skipped (no data for any mode):" footer rows that summarize_feeds
        # writes with an empty Session cell.
        if session != "(aggregate)" and session not in SESSION_ORDER:
            continue
        feed_id = row[0]
        ids = _parse_ids_cell(row[2])
        entry = feeds.setdefault(
            feed_id,
            {"aggregate": None, "sessions": {s: None for s in SESSION_ORDER}},
        )
        if session == "(aggregate)":
            entry["aggregate"] = ids
        else:
            entry["sessions"][session] = ids
    return feeds


# marketSchedule templates (America/New_York; session windows; US-equity holidays),
# used only when ADDING a missing extended-hours session. Sourced from feed 922 (AAPL).
SCHEDULE_TEMPLATES = {
    "REGULAR": (
        "America/New_York;0930-1600,0930-1600,0930-1600,0930-1600,0930-1600,C,C;"
        "0101/C,0119/C,0216/C,0403/C,0525/C,0619/C,0703/C,0907/C,1126/C,"
        "1127/0930-1300,1224/0930-1300,1225/C"
    ),
    "PRE_MARKET": (
        "America/New_York;0400-0930,0400-0930,0400-0930,0400-0930,0400-0930,C,C;"
        "0101/C,0119/C,0216/C,0403/C,0525/C,0619/C,0703/C,0907/C,1126/C,1225/C"
    ),
    "POST_MARKET": (
        "America/New_York;1600-2000,1600-2000,1600-2000,1600-2000,1600-2000,C,C;"
        "0101/C,0119/C,0216/C,0403/C,0525/C,0619/C,0703/C,0907/C,1126/C,1225/C"
    ),
    "OVER_NIGHT": (
        "America/New_York;0000-0400&2000-2400,0000-0400&2000-2400,"
        "0000-0400&2000-2400,0000-0400&2000-2400,0000-0400,C,2000-2400;"
        "0118/C,0119/2000-2400,0215/C,0216/2000-2400,0402/0000-0400,0403/C,"
        "0524/C,0525/2000-2400,0618/0000-0400,0619/C,0702/0000-0400,0703/C,"
        "0906/C,0907/2000-2400,1125/0000-0400,1126/2000-2400,1224/0000-0400,"
        "1225/C,1231/0000-0400,0101/C"
    ),
}


def _ids_inline(ids: list[int]) -> str:
    """Render an id list as an inline JSON array: '[ 1, 2, 3 ]' or '[ ]'."""
    return "[ " + ", ".join(str(i) for i in ids) + " ]" if ids else "[ ]"


def set_top_level_allowed(block: str, ids: list[int]) -> str:
    """Set the feed's top-level allowedPublisherIds.

    The top-level array is the only allowedPublisherIds that precedes
    marketSchedules, so we restrict the search to the head of the block (before
    "marketSchedules") to avoid matching a session array. The pattern spans
    multi-line arrays because [^\\]] also matches newlines. If the feed has no
    top-level allowedPublisherIds, insert one after the opening '{'.
    """
    ms = re.search(r'"marketSchedules"', block)
    head_end = ms.start() if ms else len(block)
    head = block[:head_end]
    pattern = r'"allowedPublisherIds":\s*(\[[^\]]*\]|null)'
    repl = f'"allowedPublisherIds": {_ids_inline(ids)}'
    if re.search(pattern, head):
        new_head = re.sub(pattern, repl, head, count=1)
        return new_head + block[head_end:]
    nl = block.index("\n")
    return block[:nl] + f"\n      {repl}," + block[nl:]


def _marketschedules_end(block: str) -> int:
    """Return the offset just past the marketSchedules array's closing ']', or 0."""
    ms = re.search(r'"marketSchedules":\s*\[', block)
    if not ms:
        return 0
    pos = ms.end()
    depth = 1
    in_str = False
    while pos < len(block) and depth > 0:
        c = block[pos]
        if c == '"' and (pos == 0 or block[pos - 1] != "\\"):
            in_str = not in_str
        elif not in_str:
            if c == "[":
                depth += 1
            elif c == "]":
                depth -= 1
        pos += 1
    return pos


def set_top_level_min_publishers(block: str, n: int) -> str:
    """Set the feed's top-level minPublishers.

    after.json lists marketSchedules (with their own minPublishers) BEFORE the
    top-level minPublishers, so we search only the region after the
    marketSchedules array closes. Falls back to the first match when the feed
    has no marketSchedules.
    """
    search_from = _marketschedules_end(block)
    pat = r'"minPublishers":\s*\d+'
    m = re.search(pat, block[search_from:])
    if not m:
        return re.sub(pat, f'"minPublishers": {n}', block, count=1)
    s = search_from + m.start()
    e = search_from + m.end()
    return block[:s] + f'"minPublishers": {n}' + block[e:]


def overwrite_session(block: str, session: str, ids: list[int]) -> str:
    """Within a feed block, set a session's allowedPublisherIds + minPublishers."""
    bounds = find_session_block(block, session)
    if bounds is None:
        return block
    s, e = bounds
    sblock = block[s:e]
    min_pub = get_min_publishers(session, len(ids))

    pub_pat = r'"allowedPublisherIds":\s*(\[[^\]]*\]|null)'
    if re.search(pub_pat, sblock):
        sblock = re.sub(
            pub_pat, f'"allowedPublisherIds": {_ids_inline(ids)}', sblock, count=1
        )
    min_pat = r'"minPublishers":\s*\d+'
    if re.search(min_pat, sblock):
        sblock = re.sub(min_pat, f'"minPublishers": {min_pub}', sblock, count=1)
    return block[:s] + sblock + block[e:]


def _detect_session_indent(block: str) -> str:
    """Return the leading whitespace of the first session entry's '{', or 8 spaces."""
    m = re.search(r"\n(\s*)\{", block[block.find('"marketSchedules"') :])
    return m.group(1) if m else "        "


def add_session(block: str, session: str, ids: list[int], benchmark_mapping) -> str:
    """Insert a new session entry before the closing ']' of marketSchedules.

    benchmark_mapping is the dict copied from the feed's REGULAR session (or None).
    """
    base_indent = _detect_session_indent(block)
    entry: dict = {"allowedPublisherIds": ids}
    if benchmark_mapping is not None:
        entry["benchmarkMapping"] = benchmark_mapping
    entry["marketSchedule"] = SCHEDULE_TEMPLATES[session]
    entry["minPublishers"] = get_min_publishers(session, len(ids))
    entry["session"] = session

    text = json.dumps(entry, indent=2)
    entry_text = "\n".join(base_indent + ln for ln in text.split("\n"))

    ms_end = _marketschedules_end(block)
    if ms_end == 0:
        return block
    closing_bracket = ms_end - 1  # position of the array's ']'

    # Walk back to the last non-whitespace char before ']'. For a non-empty
    # array that is the previous entry's '}' (prepend a comma); for an empty
    # array it is the opening '[' (no comma, else invalid JSON).
    p = closing_bracket - 1
    while p >= 0 and block[p] in (" ", "\n", "\t", "\r"):
        p -= 1
    sep = "\n" if block[p] == "[" else ",\n"
    return block[: p + 1] + sep + entry_text + block[p + 1 :]


def _regular_benchmark_mapping(feed: dict):
    """Return the benchmarkMapping dict from the feed's REGULAR session, or None."""
    for s in feed.get("marketSchedules", []):
        if s.get("session") == "REGULAR":
            return s.get("benchmarkMapping")
    return None


def apply_summary_to_config(
    raw: str, summary: dict[int, dict], log=None
) -> tuple[str, dict]:
    """Apply the parsed summary to the raw config text.

    Returns (new_raw, stats). `log` is an optional callable(str) for per-feed
    lines; defaults to a no-op. Implements the spec decision matrix.
    """
    if log is None:
        log = lambda _msg: None  # noqa: E731

    data = json.loads(raw)
    feed_index = {f["feedId"]: f for f in data["feeds"]}

    stats = {
        "promoted": 0,
        "sessions_added": 0,
        "skipped_no_data": 0,
        "skipped_no_publishers": 0,
        "skipped_state": 0,
        "not_found": [],
        "filtered_any": False,
    }

    for feed_id, fa in summary.items():
        if not fa["aggregate"]:
            stats["skipped_no_data"] += 1
            log(f"  SKIP (no data): feedId={feed_id}")
            continue

        feed = feed_index.get(feed_id)
        if feed is None:
            stats["not_found"].append(feed_id)
            log(f"  WARNING (not found): feedId={feed_id}")
            continue

        state = feed.get("state")
        if state not in ("COMING_SOON", "STABLE"):
            stats["skipped_state"] += 1
            log(f"  SKIP (state={state}): feedId={feed_id}")
            continue

        bounds = find_feed_block(raw, feed_id)
        if bounds is None:
            stats["not_found"].append(feed_id)
            log(f"  WARNING (block not found): feedId={feed_id}")
            continue

        start, end = bounds
        block = raw[start:end]
        existing_sessions = {s.get("session") for s in feed.get("marketSchedules", [])}
        bench = _regular_benchmark_mapping(feed)

        if state == "COMING_SOON":
            top_union: set[int] = set()
            for session in SESSION_ORDER:
                raw_ids = fa["sessions"].get(session)
                if not raw_ids:
                    continue
                kept, removed = filter_publishers(raw_ids)
                if removed:
                    stats["filtered_any"] = True
                    log(f"    filtered {removed} from {feed_id}/{session}")
                if not kept:
                    continue
                top_union.update(kept)
                if session in existing_sessions:
                    block = overwrite_session(block, session, kept)
                else:
                    block = add_session(block, session, kept, bench)
                    stats["sessions_added"] += 1
            if not top_union:
                # Aggregate had ids but everything filtered out (e.g. only
                # excluded/Lazer publishers). Do NOT promote to STABLE with an
                # empty allow-list; leave the feed COMING_SOON. No edits were
                # written to `block` in this case (every session was skipped).
                stats["skipped_no_publishers"] += 1
                log(f"  SKIP (no publishers after filter): feedId={feed_id}")
                continue
            block = re.sub(
                r'"state":\s*"COMING_SOON"', '"state": "STABLE"', block, count=1
            )
            block = set_top_level_allowed(block, sorted(top_union))
            block = set_top_level_min_publishers(block, 1)
            stats["promoted"] += 1
            log(f"  PROMOTE: feedId={feed_id} -> STABLE, top={sorted(top_union)}")
        else:  # STABLE — additive only
            added: set[int] = set()
            for session in SESSION_ORDER:
                raw_ids = fa["sessions"].get(session)
                if not raw_ids:
                    continue
                if session in existing_sessions:
                    log(f"  SKIP (live): feedId={feed_id}/{session}")
                    continue
                kept, removed = filter_publishers(raw_ids)
                if removed:
                    stats["filtered_any"] = True
                    log(f"    filtered {removed} from {feed_id}/{session}")
                if not kept:
                    continue
                block = add_session(block, session, kept, bench)
                added.update(kept)
                stats["sessions_added"] += 1
                log(f"  ADD-SESSION: feedId={feed_id}/{session}={kept}")
            if added:
                existing_top = feed.get("allowedPublisherIds") or []
                new_top = sorted(set(existing_top) | added)
                block = set_top_level_allowed(block, new_top)

        raw = raw[:start] + block + raw[end:]

    return raw, stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Apply a dq_summary 'allowed' sheet to after.json."
    )
    parser.add_argument(
        "--xlsx", required=True, help="dq_summary_<cluster>_<date>.xlsx"
    )
    parser.add_argument(
        "--config", required=True, help="after.json / after_1.json config file"
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Preview changes; write nothing."
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    config_path = Path(args.config)
    if not xlsx_path.exists():
        print(f"ERROR: workbook not found: {xlsx_path}")
        sys.exit(1)
    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}")
        sys.exit(1)

    print(f"Reading allowed sheet from {xlsx_path}")
    summary = parse_allowed_sheet(xlsx_path)
    print(f"Found {len(summary)} feeds in the allowed sheet")

    if args.dry_run:
        print("\n=== DRY RUN (no files will be modified) ===\n")

    raw = config_path.read_text()
    new_raw, stats = apply_summary_to_config(raw, summary, log=print)

    changed = stats["promoted"] + stats["sessions_added"]
    if not args.dry_run and changed > 0:
        backup = str(config_path) + ".bak"
        shutil.copy2(config_path, backup)
        config_path.write_text(new_raw)
        print(f"\nBackup saved to {backup}")

    print(f"\n{'=' * 50}\nSUMMARY\n{'=' * 50}")
    print(f"  Feeds promoted (COMING_SOON->STABLE): {stats['promoted']}")
    print(f"  Sessions added:                       {stats['sessions_added']}")
    print(f"  Skipped (no data):                    {stats['skipped_no_data']}")
    print(f"  Skipped (no publishers after filter): {stats['skipped_no_publishers']}")
    print(f"  Skipped (other state):                {stats['skipped_state']}")
    print(f"  Not found in config:                  {len(stats['not_found'])}")
    if stats["not_found"]:
        print(f"  Missing feed IDs: {stats['not_found']}")
    if stats["filtered_any"]:
        print("  NOTE: some Lazer/zero publishers were filtered (see lines above).")


if __name__ == "__main__":
    main()
