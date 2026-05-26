"""Unit + integration tests for apply_allowed_to_config."""
import json
from pathlib import Path

import openpyxl
import pytest

from lazer_dq.apply_allowed_to_config import parse_allowed_sheet


def _write_allowed_workbook(path: Path, rows: list[tuple]) -> None:
    """rows: list of (feed_id, session, allowed_cell, note). Builds a 2-sheet
    workbook matching summarize_feeds output (rankings sheet present but empty)."""
    wb = openpyxl.Workbook()
    wb.active.title = "rankings"  # present but unused by the reader
    ws = wb.create_sheet("allowed")
    ws.cell(row=1, column=1, value="Allowed Publishers — test — 2026-05-20")
    for i, h in enumerate(["Feed ID", "Session", "allowedPublisherIds", "Notes"], 1):
        ws.cell(row=2, column=i, value=h)
    r = 3
    for feed_id, session, allowed_cell, note in rows:
        ws.cell(row=r, column=1, value=feed_id)
        ws.cell(row=r, column=2, value=session)
        ws.cell(row=r, column=3, value=allowed_cell)
        ws.cell(row=r, column=4, value=note)
        r += 1
    wb.save(path)


def _agg(ids):
    return '"allowedPublisherIds": [ ' + ", ".join(str(i) for i in ids) + " ],"


def test_parse_allowed_sheet_reads_lists_and_no_data(tmp_path):
    xlsx = tmp_path / "dq.xlsx"
    _write_allowed_workbook(
        xlsx,
        [
            (100, "(aggregate)", _agg([24, 35, 42]), None),
            (100, "REGULAR", _agg([24, 35, 42]), "0 passed + 3 top-up (≤2×)"),
            (100, "PRE_MARKET", "(no data)", "mode missing for 2026-05-20"),
            (100, "POST_MARKET", "(no data)", "mode missing for 2026-05-20"),
            (100, "OVER_NIGHT", "(no data)", "mode missing for 2026-05-20"),
            (None, None, None, None),  # divider
            (200, "(aggregate)", "(no data)", "all sessions empty"),
            (200, "REGULAR", "(no data)", "mode missing for 2026-05-20"),
        ],
    )

    result = parse_allowed_sheet(xlsx)

    assert set(result.keys()) == {100, 200}
    assert result[100]["aggregate"] == [24, 35, 42]
    assert result[100]["sessions"]["REGULAR"] == [24, 35, 42]
    assert result[100]["sessions"]["PRE_MARKET"] is None
    assert result[200]["aggregate"] is None
    assert result[200]["sessions"]["REGULAR"] is None


def test_parse_allowed_sheet_ignores_skipped_footer_rows(tmp_path):
    xlsx = tmp_path / "dq_footer.xlsx"
    _write_allowed_workbook(
        xlsx,
        [
            (100, "(aggregate)", _agg([24, 35]), None),
            (100, "REGULAR", _agg([24, 35]), None),
            (None, None, None, None),  # divider
            # Simulated "Feeds skipped" footer: bare integer feed id, no session.
            (777, None, None, None),
        ],
    )
    result = parse_allowed_sheet(xlsx)
    assert set(result.keys()) == {100}  # 777 footer row must NOT become a feed
    assert result[100]["aggregate"] == [24, 35]


from lazer_dq.apply_allowed_to_config import filter_publishers, get_min_publishers


def test_filter_publishers_strips_zero_and_lazer():
    kept, removed = filter_publishers([0, 1, 9, 13, 15, 24, 35, 42])
    assert kept == [24, 35, 42]
    assert removed == [0, 1, 9, 13, 15]


def test_filter_publishers_keeps_sorted_unique():
    kept, removed = filter_publishers([42, 24, 24, 35])
    assert kept == [24, 35, 42]
    assert removed == []


def test_get_min_publishers_defaults():
    assert get_min_publishers("REGULAR", 10) == 3
    assert get_min_publishers("PRE_MARKET", 10) == 2
    assert get_min_publishers("POST_MARKET", 10) == 2
    assert get_min_publishers("OVER_NIGHT", 10) == 1


def test_get_min_publishers_regular_low_count_rule():
    assert get_min_publishers("REGULAR", 5) == 2
    assert get_min_publishers("REGULAR", 1) == 2
    assert get_min_publishers("REGULAR", 6) == 3


from lazer_dq.apply_allowed_to_config import (
    set_top_level_allowed,
    set_top_level_min_publishers,
    overwrite_session,
    add_session,
    SCHEDULE_TEMPLATES,
)


def test_set_top_level_allowed_replaces_array_before_marketschedules():
    block = (
        '{\n      "allowedPublisherIds": [\n        1,\n        2\n      ],\n'
        '      "marketSchedules": [ {"allowedPublisherIds": [9], '
        '"session": "REGULAR"} ]\n}'
    )
    out = set_top_level_allowed(block, [24, 35])
    # Top-level array (before marketSchedules) replaced; session array untouched.
    assert '"allowedPublisherIds": [ 24, 35 ]' in out
    assert '"allowedPublisherIds": [9]' in out
    assert out.index("[ 24, 35 ]") < out.index("[9]")


def test_set_top_level_min_publishers_targets_field_after_marketschedules():
    # Mirrors after.json: a session minPublishers appears BEFORE the top-level one.
    block = (
        '{\n      "allowedPublisherIds": [ 1 ],\n'
        '      "marketSchedules": [ {\n'
        '          "minPublishers": 3,\n'
        '          "session": "REGULAR"\n'
        "        } ],\n"
        '      "minPublishers": 3,\n'
        '      "state": "STABLE"\n}'
    )
    out = set_top_level_min_publishers(block, 1)
    data = json.loads(out)
    assert data["minPublishers"] == 1  # top-level changed
    assert data["marketSchedules"][0]["minPublishers"] == 3  # session untouched


def test_overwrite_session_replaces_ids_and_minpub():
    block = (
        '{ "marketSchedules": [ {\n'
        '          "allowedPublisherIds": [ 1, 2, 3 ],\n'
        '          "minPublishers": 3,\n'
        '          "session": "REGULAR"\n'
        "        } ] }"
    )
    out = overwrite_session(block, "REGULAR", [24, 35, 42])
    assert '"allowedPublisherIds": [ 24, 35, 42 ]' in out
    assert '"minPublishers": 2' in out  # 3 publishers => REGULAR low-count => 2
    assert '"session": "REGULAR"' in out


def test_overwrite_session_handles_null_array():
    block = (
        '{ "marketSchedules": [ {\n'
        '          "allowedPublisherIds": null,\n'
        '          "minPublishers": 3,\n'
        '          "session": "PRE_MARKET"\n'
        "        } ] }"
    )
    out = overwrite_session(block, "PRE_MARKET", [24, 35])
    assert '"allowedPublisherIds": [ 24, 35 ]' in out
    assert "null" not in out
    assert '"minPublishers": 2' in out


def test_overwrite_session_inserts_when_fields_absent():
    # Real after.json shape: a COMING_SOON REGULAR entry with NO
    # allowedPublisherIds and NO minPublishers key.
    block = (
        '{ "marketSchedules": [\n'
        "        {\n"
        '          "benchmarkMapping": {"datascope_ric": {}},\n'
        '          "marketSchedule": "X",\n'
        '          "session": "REGULAR"\n'
        "        }\n"
        "      ] }"
    )
    out = overwrite_session(block, "REGULAR", [24, 35, 42])
    data = json.loads(out)  # must be valid JSON
    reg = data["marketSchedules"][0]
    assert reg["allowedPublisherIds"] == [24, 35, 42]
    assert reg["minPublishers"] == 2  # 3 pubs => REGULAR low-count
    assert reg["benchmarkMapping"] == {"datascope_ric": {}}  # preserved
    # Inserted minPublishers sits between marketSchedule and session (canonical
    # order); allowedPublisherIds leads. Key order is preserved by json.loads.
    assert list(reg.keys()) == [
        "allowedPublisherIds",
        "benchmarkMapping",
        "marketSchedule",
        "minPublishers",
        "session",
    ]


def test_add_session_inserts_entry_with_benchmark_mapping():
    block = (
        '{ "marketSchedules": [\n'
        "        {\n"
        '          "allowedPublisherIds": [ 11 ],\n'
        '          "marketSchedule": "X",\n'
        '          "minPublishers": 3,\n'
        '          "session": "REGULAR"\n'
        "        }\n"
        "      ]\n}"
    )
    bench = {"datascope_ric": {"identifiers": [{"identifier": "AAPL.O"}]}}
    out = add_session(block, "PRE_MARKET", [24, 35], bench)
    # Still valid JSON after the insert.
    data = json.loads(out)
    sessions = {s["session"]: s for s in data["marketSchedules"]}
    assert set(sessions) == {"REGULAR", "PRE_MARKET"}
    pre = sessions["PRE_MARKET"]
    assert pre["allowedPublisherIds"] == [24, 35]
    assert pre["minPublishers"] == 2  # PRE_MARKET default
    assert pre["benchmarkMapping"] == bench
    assert pre["marketSchedule"] == SCHEDULE_TEMPLATES["PRE_MARKET"]
    # REGULAR untouched.
    assert sessions["REGULAR"]["allowedPublisherIds"] == [11]


def test_add_session_into_empty_marketschedules():
    block = '{ "marketSchedules": [] }'
    out = add_session(block, "REGULAR", [24, 35], None)
    data = json.loads(out)  # must be valid JSON
    assert len(data["marketSchedules"]) == 1
    sess = data["marketSchedules"][0]
    assert sess["session"] == "REGULAR"
    assert sess["allowedPublisherIds"] == [24, 35]
    assert "benchmarkMapping" not in sess  # None mapping omitted


from lazer_dq.apply_allowed_to_config import apply_summary_to_config

_BENCH = {"datascope_ric": {"identifiers": [{"identifier": "AAPL.O"}]}}


def _config_with(feeds: list[dict]) -> str:
    """Serialize a minimal config the way after.json is laid out (indent=2)."""
    return json.dumps({"feeds": feeds}, indent=2)


def _feed(feed_id, state, sessions, top=None):
    """sessions: list of (name, allowed_or_None). REGULAR carries benchmarkMapping."""
    ms = []
    for name, allowed in sessions:
        entry = {
            "allowedPublisherIds": allowed,
            "benchmarkMapping": _BENCH,
            "marketSchedule": "TPL",
            "minPublishers": 3,
            "session": name,
        }
        ms.append(entry)
    feed = {
        "allowedPublisherIds": top if top is not None else [],
        "feedId": feed_id,
        "marketSchedules": ms,
        "minPublishers": 3,
        "state": state,
        "symbol": f"S{feed_id}",
    }
    return feed


def test_apply_promotes_coming_soon_regular_only():
    raw = _config_with(
        [_feed(100, "COMING_SOON", [("REGULAR", [1, 2, 3])], top=[1, 2, 3])]
    )
    summary = {
        100: {
            "aggregate": [24, 35, 42],
            "sessions": {
                "REGULAR": [24, 35, 42],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    data = json.loads(out)
    feed = {f["feedId"]: f for f in data["feeds"]}[100]

    assert feed["state"] == "STABLE"
    assert feed["minPublishers"] == 2  # top-level set to 2 on promotion
    assert feed["allowedPublisherIds"] == [24, 35, 42]
    reg = feed["marketSchedules"][0]
    assert reg["allowedPublisherIds"] == [24, 35, 42]
    assert reg["minPublishers"] == 2  # 3 pubs => REGULAR low-count
    assert stats["promoted"] == 1


def test_apply_adds_missing_session_to_stable_feed():
    raw = _config_with([_feed(200, "STABLE", [("REGULAR", [11, 12])], top=[11, 12])])
    summary = {
        200: {
            "aggregate": [24, 35],
            "sessions": {
                "REGULAR": [11, 12],
                "PRE_MARKET": [24, 35],
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    data = json.loads(out)
    feed = {f["feedId"]: f for f in data["feeds"]}[200]
    sess = {s["session"]: s for s in feed["marketSchedules"]}

    assert feed["state"] == "STABLE"  # unchanged
    assert sess["REGULAR"]["allowedPublisherIds"] == [11, 12]  # live, untouched
    assert sess["PRE_MARKET"]["allowedPublisherIds"] == [24, 35]  # added
    assert sess["PRE_MARKET"]["benchmarkMapping"] == _BENCH  # copied from REGULAR
    assert feed["allowedPublisherIds"] == [11, 12, 24, 35]  # folded union
    assert feed["minPublishers"] == 3  # top-level untouched on STABLE
    assert stats["sessions_added"] == 1
    assert stats["skipped_stable_no_change"] == 0  # a session WAS added


def test_apply_leaves_existing_stable_session_untouched():
    raw = _config_with(
        [_feed(300, "STABLE", [("REGULAR", [11]), ("PRE_MARKET", [99])], top=[11, 99])]
    )
    summary = {
        300: {
            "aggregate": [24],
            "sessions": {
                "REGULAR": [11],
                "PRE_MARKET": [24],
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    data = json.loads(out)
    feed = {f["feedId"]: f for f in data["feeds"]}[300]
    sess = {s["session"]: s for s in feed["marketSchedules"]}

    assert sess["PRE_MARKET"]["allowedPublisherIds"] == [99]  # NOT overwritten
    assert stats["sessions_added"] == 0
    assert feed == {f["feedId"]: f for f in json.loads(raw)["feeds"]}[300]  # untouched
    assert stats["skipped_stable_no_change"] == 1  # STABLE, nothing new to add


def test_apply_skips_no_data_feed():
    raw = _config_with([_feed(400, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        400: {
            "aggregate": None,
            "sessions": {
                s: None for s in ["REGULAR", "PRE_MARKET", "POST_MARKET", "OVER_NIGHT"]
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    assert out == raw  # nothing changed
    assert stats["skipped_no_data"] == 1


def test_apply_warns_on_missing_feed():
    raw = _config_with([_feed(500, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        999: {
            "aggregate": [24],
            "sessions": {
                "REGULAR": [24],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    assert out == raw
    assert stats["not_found"] == [999]


def test_apply_filters_lazer_and_warns():
    raw = _config_with([_feed(600, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        600: {
            "aggregate": [1, 9, 24, 35, 42],
            "sessions": {
                "REGULAR": [1, 9, 24, 35, 42],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    data = json.loads(out)
    feed = {f["feedId"]: f for f in data["feeds"]}[600]
    assert feed["marketSchedules"][0]["allowedPublisherIds"] == [24, 35, 42]
    assert feed["allowedPublisherIds"] == [24, 35, 42]
    assert stats["filtered_any"] is True


def test_apply_does_not_promote_when_all_publishers_filtered():
    raw = _config_with([_feed(700, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        700: {
            "aggregate": [1, 9, 13],
            "sessions": {
                "REGULAR": [1, 9, 13],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    data = json.loads(out)
    feed = {f["feedId"]: f for f in data["feeds"]}[700]

    assert feed["state"] == "COMING_SOON"  # NOT promoted (0 survive filtering)
    assert stats["promoted"] == 0
    assert stats["skipped_too_few_publishers"] == 1
    assert stats["filtered_any"] is True


def test_apply_does_not_promote_fewer_than_three_publishers():
    # 2 publishers survive filtering -> below the redundancy gate -> not promoted.
    raw = _config_with([_feed(800, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        800: {
            "aggregate": [24, 35],
            "sessions": {
                "REGULAR": [24, 35],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    assert out == raw  # block untouched; no partial edits
    feed = {f["feedId"]: f for f in json.loads(out)["feeds"]}[800]
    assert feed["state"] == "COMING_SOON"  # NOT promoted
    assert stats["promoted"] == 0
    assert stats["sessions_added"] == 0
    assert stats["skipped_too_few_publishers"] == 1


def test_apply_promotes_with_exactly_three_publishers():
    # Exactly 3 survivors clears the gate.
    raw = _config_with([_feed(810, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        810: {
            "aggregate": [24, 35, 42],
            "sessions": {
                "REGULAR": [24, 35, 42],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary)
    feed = {f["feedId"]: f for f in json.loads(out)["feeds"]}[810]
    assert feed["state"] == "STABLE"
    assert feed["allowedPublisherIds"] == [24, 35, 42]
    assert stats["promoted"] == 1


def test_apply_min_promote_publishers_param_lowers_gate():
    # A 2-publisher feed is skipped at the default gate (3) but promotes when
    # the caller lowers min_promote_publishers to 2.
    raw = _config_with([_feed(820, "COMING_SOON", [("REGULAR", [1])], top=[1])])
    summary = {
        820: {
            "aggregate": [24, 35],
            "sessions": {
                "REGULAR": [24, 35],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    # Default gate (3): not promoted.
    out_default, stats_default = apply_summary_to_config(raw, summary)
    assert stats_default["promoted"] == 0
    assert stats_default["skipped_too_few_publishers"] == 1

    # Lowered gate (2): promoted.
    out_low, stats_low = apply_summary_to_config(raw, summary, min_promote_publishers=2)
    feed = {f["feedId"]: f for f in json.loads(out_low)["feeds"]}[820]
    assert feed["state"] == "STABLE"
    assert feed["allowedPublisherIds"] == [24, 35]
    assert stats_low["promoted"] == 1


def test_apply_write_session_fields_false_sets_top_level_only():
    # hk-equities shape: a COMING_SOON feed whose REGULAR entry has NO
    # allowedPublisherIds and NO minPublishers (just benchmarkMapping/schedule).
    feed = {
        "allowedPublisherIds": [1, 3, 5],  # COMING_SOON placeholder
        "feedId": 884,
        "marketSchedules": [
            {
                "benchmarkMapping": _BENCH,
                "marketSchedule": "Asia/Hong_Kong;0930-1200,C",
                "session": "REGULAR",
            }
        ],
        "minPublishers": 3,
        "state": "COMING_SOON",
        "symbol": "S884",
    }
    raw = json.dumps({"feeds": [feed]}, indent=2)
    summary = {
        884: {
            "aggregate": [41, 69],
            "sessions": {
                "REGULAR": [41, 69],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(
        raw, summary, min_promote_publishers=2, write_session_fields=False
    )
    f = {x["feedId"]: x for x in json.loads(out)["feeds"]}[884]

    assert f["state"] == "STABLE"
    assert f["allowedPublisherIds"] == [41, 69]  # top-level set
    assert f["minPublishers"] == 2  # top-level set
    reg = f["marketSchedules"][0]
    # REGULAR entry left exactly as-is — no session-level fields added.
    assert set(reg.keys()) == {"benchmarkMapping", "marketSchedule", "session"}
    assert stats["promoted"] == 1
    assert stats["sessions_added"] == 0


import subprocess
import sys


def _real_workbook(tmp_path):
    xlsx = tmp_path / "dq_summary_test_2026-05-20.xlsx"
    _write_allowed_workbook(
        xlsx,
        [
            (100, "(aggregate)", _agg([24, 35, 42]), None),
            (100, "REGULAR", _agg([24, 35, 42]), "0 passed + 3 top-up (≤2×)"),
            (100, "PRE_MARKET", "(no data)", "mode missing"),
            (100, "POST_MARKET", "(no data)", "mode missing"),
            (100, "OVER_NIGHT", "(no data)", "mode missing"),
        ],
    )
    return xlsx


def _real_config(tmp_path):
    cfg = tmp_path / "after_test.json"
    cfg.write_text(
        _config_with(
            [_feed(100, "COMING_SOON", [("REGULAR", [1, 2, 3])], top=[1, 2, 3])]
        )
    )
    return cfg


def test_cli_dry_run_writes_nothing(tmp_path):
    xlsx = _real_workbook(tmp_path)
    cfg = _real_config(tmp_path)
    before = cfg.read_text()

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "lazer_dq.apply_allowed_to_config",
            "--xlsx",
            str(xlsx),
            "--config",
            str(cfg),
            "--dry-run",
        ],
        capture_output=True,
        text=True,
        cwd=str(Path(__file__).resolve().parents[2]),
    )
    assert result.returncode == 0, result.stderr
    assert "DRY RUN" in result.stdout
    assert cfg.read_text() == before  # unchanged
    assert not (tmp_path / "after_test.json.bak").exists()


def test_cli_real_run_writes_and_backs_up(tmp_path):
    xlsx = _real_workbook(tmp_path)
    cfg = _real_config(tmp_path)

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "lazer_dq.apply_allowed_to_config",
            "--xlsx",
            str(xlsx),
            "--config",
            str(cfg),
        ],
        capture_output=True,
        text=True,
        cwd=str(Path(__file__).resolve().parents[2]),
    )
    assert result.returncode == 0, result.stderr
    assert (cfg.parent / "after_test.json.bak").exists()
    data = json.loads(cfg.read_text())
    feed = {f["feedId"]: f for f in data["feeds"]}[100]
    assert feed["state"] == "STABLE"
    assert feed["allowedPublisherIds"] == [24, 35, 42]


from lazer_dq.apply_allowed_to_config import remove_session


def _ms_block(sessions):
    """A feed block whose marketSchedules holds one entry per session name."""
    entries = [
        "        {\n"
        '          "allowedPublisherIds": null,\n'
        '          "minPublishers": 1,\n'
        f'          "session": "{s}"\n'
        "        }"
        for s in sessions
    ]
    return '{\n      "marketSchedules": [\n' + ",\n".join(entries) + "\n      ]\n}"


def test_remove_session_first():
    out = remove_session(_ms_block(["REGULAR", "PRE_MARKET", "OVER_NIGHT"]), "REGULAR")
    assert [s["session"] for s in json.loads(out)["marketSchedules"]] == [
        "PRE_MARKET",
        "OVER_NIGHT",
    ]


def test_remove_session_middle():
    out = remove_session(
        _ms_block(["REGULAR", "PRE_MARKET", "OVER_NIGHT"]), "PRE_MARKET"
    )
    assert [s["session"] for s in json.loads(out)["marketSchedules"]] == [
        "REGULAR",
        "OVER_NIGHT",
    ]


def test_remove_session_last():
    out = remove_session(
        _ms_block(["REGULAR", "PRE_MARKET", "OVER_NIGHT"]), "OVER_NIGHT"
    )
    assert [s["session"] for s in json.loads(out)["marketSchedules"]] == [
        "REGULAR",
        "PRE_MARKET",
    ]


def test_remove_session_absent_is_noop():
    block = _ms_block(["REGULAR"])
    assert remove_session(block, "PRE_MARKET") == block


def test_remove_session_sequential_down_to_one():
    block = _ms_block(["REGULAR", "PRE_MARKET", "POST_MARKET", "OVER_NIGHT"])
    for s in ["PRE_MARKET", "POST_MARKET", "OVER_NIGHT"]:
        block = remove_session(block, s)
    assert [s["session"] for s in json.loads(block)["marketSchedules"]] == ["REGULAR"]


def test_apply_promotion_drops_sessions_without_publishers():
    # us-equities COMING_SOON feed with all 4 sessions; only REGULAR has publishers
    # in the summary (the real feed 2300 / 2026-05-20 shape).
    feed = {
        "allowedPublisherIds": [1, 2, 3],
        "feedId": 2300,
        "marketSchedules": [
            {
                "allowedPublisherIds": [1],
                "benchmarkMapping": _BENCH,
                "minPublishers": 3,
                "session": "REGULAR",
            },
            {"allowedPublisherIds": None, "minPublishers": 1, "session": "PRE_MARKET"},
            {"allowedPublisherIds": None, "minPublishers": 1, "session": "POST_MARKET"},
            {"allowedPublisherIds": None, "minPublishers": 1, "session": "OVER_NIGHT"},
        ],
        "minPublishers": 3,
        "state": "COMING_SOON",
        "symbol": "S2300",
    }
    raw = json.dumps({"feeds": [feed]}, indent=2)
    summary = {
        2300: {
            "aggregate": [19, 41],
            "sessions": {
                "REGULAR": [19, 41],
                "PRE_MARKET": None,
                "POST_MARKET": None,
                "OVER_NIGHT": None,
            },
        }
    }

    out, stats = apply_summary_to_config(raw, summary, min_promote_publishers=2)
    f = {x["feedId"]: x for x in json.loads(out)["feeds"]}[2300]

    assert f["state"] == "STABLE"
    # PRE/POST/OVERNIGHT dropped; only the priced REGULAR session remains.
    assert [s["session"] for s in f["marketSchedules"]] == ["REGULAR"]
    assert f["marketSchedules"][0]["allowedPublisherIds"] == [19, 41]
    assert f["allowedPublisherIds"] == [19, 41]
    assert stats["sessions_removed"] == 3
    assert stats["promoted"] == 1
