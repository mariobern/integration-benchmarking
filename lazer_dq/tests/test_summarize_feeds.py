"""Unit + integration tests for summarize_feeds."""
from pathlib import Path

import pytest

from lazer_dq.summarize_feeds import (
    load_excluded_publishers,
)


# ---------- load_excluded_publishers ----------


def _write_publishers_md(tmp_path: Path, body: str) -> Path:
    p = tmp_path / "publishers.md"
    p.write_text(body)
    return p


def test_load_excluded_publishers_extracts_dot_test_and_zero(tmp_path):
    md = _write_publishers_md(
        tmp_path,
        """\
# Publisher IDs and Names
| ID  | Name                  | Active |
| --- | --------------------- | ------ |
| 1   | Lazer.Binance         | Yes    |
| 23  | LoTech.Test           | Yes    |
| 25  | CharlesworthResearch.Test | Yes |
| 26  | CharlesworthResearch.Production | Yes |
""",
    )
    assert load_excluded_publishers(md) == {0, 23, 25}


def test_load_excluded_publishers_always_includes_zero_even_if_empty_md(tmp_path):
    md = _write_publishers_md(tmp_path, "# empty\n")
    assert load_excluded_publishers(md) == {0}


def test_load_excluded_publishers_handles_malformed_row(tmp_path):
    md = _write_publishers_md(
        tmp_path,
        """\
| ID  | Name        | Active |
| --- | ----------- | ------ |
| abc | Bad.Test    | Yes    |
| 27  | MEMX.Test   | Yes    |
""",
    )
    # Malformed ID row skipped, valid one parsed.
    assert load_excluded_publishers(md) == {0, 27}


def test_load_excluded_publishers_ignores_production_publishers(tmp_path):
    md = _write_publishers_md(
        tmp_path,
        """\
| ID  | Name              | Active |
| --- | ----------------- | ------ |
| 1   | Lazer.Binance     | Yes    |
| 2   | Jump.Production   | Yes    |
""",
    )
    assert load_excluded_publishers(md) == {0}


from lazer_dq.summarize_feeds import discover_feeds


# ---------- discover_feeds ----------


def test_discover_feeds_returns_distinct_feed_ids_from_csv(tmp_path):
    csv = tmp_path / "input.csv"
    csv.write_text(
        "1021, 2026-05-06, us-equities-pre\n"
        "1060, 2026-05-06, us-equities-pre\n"
        "1021, 2026-05-06, us-equities-post\n"  # duplicate feed_id
        "922, 2026-05-06, us-equities\n"
    )
    assert discover_feeds(csv) == [1021, 1060, 922]


def test_discover_feeds_skips_malformed_rows(tmp_path, capsys):
    csv = tmp_path / "input.csv"
    csv.write_text(
        "1021, 2026-05-06, us-equities-pre\n"
        "\n"  # blank line
        ", , \n"  # empty fields
        "abc, 2026-05-06, us-equities\n"  # non-numeric feed_id
        "1060, 2026-05-06, us-equities\n"
    )
    assert discover_feeds(csv) == [1021, 1060]
    out = capsys.readouterr().out
    assert "abc" in out  # warning emitted


def test_discover_feeds_preserves_first_seen_order(tmp_path):
    csv = tmp_path / "input.csv"
    csv.write_text("3, x, y\n1, x, y\n2, x, y\n3, x, y\n1, x, y\n")
    assert discover_feeds(csv) == [3, 1, 2]


from lazer_dq.summarize_feeds import load_stats


# ---------- load_stats ----------

STATS_HEADER = (
    "feed_id,publisher_id,n_observations,mean_diff,std_diff,mean_pct_diff,"
    "std_pct_diff,rmse,nrmse,rmse_over_spread,mae,t_statistic,t_pvalue,"
    "wilcoxon_statistic,wilcoxon_pvalue,normality_pvalue,hit_rate_0.1pct,"
    "mean_abs_z_score,pass_fail\n"
)


def _write_stats_csv(
    reports_dir: Path, cluster, mode, feed_id, date, body_rows, header=None
):
    """Build dq_reports/<cluster>/<mode>/<feed_id>/<date>/stats.csv.

    `header` defaults to the canonical STATS_HEADER. Tests that need a
    minimal/custom column set (e.g. just the columns the code actually reads)
    can pass their own header line.
    """
    if header is None:
        header = STATS_HEADER
    p = reports_dir / cluster / mode / str(feed_id) / date / "stats.csv"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(header + "".join(body_rows))
    return p


def test_load_stats_returns_none_for_missing_file(tmp_path):
    assert load_stats(tmp_path, "lazer-prod", "us-equities", 1021, "2026-05-06") is None


def test_load_stats_parses_real_csv_format(tmp_path):
    _write_stats_csv(
        tmp_path,
        "lazer-prod",
        "us-equities-post",
        1021,
        "2026-05-06",
        [
            "1021,11,22218,-0.05,0.08,-0.01,0.02,0.0932,0.51,0.0185,0.07,-84,0,75,0,0,100.0,0.96,fail\n"
        ],
    )
    rows = load_stats(tmp_path, "lazer-prod", "us-equities-post", 1021, "2026-05-06")
    assert rows is not None
    assert len(rows) == 1
    assert rows[0]["publisher_id"] == "11"
    assert rows[0]["rmse_over_spread"] == "0.0185"
    assert rows[0]["hit_rate_0.1pct"] == "100.0"


from lazer_dq.summarize_feeds import rank_top_n


# ---------- rank_top_n ----------


def _stat(publisher_id, ros, hit=80.0, n_obs=10000):
    """Helper: minimal stats.csv-style dict."""
    return {
        "publisher_id": str(publisher_id),
        "rmse_over_spread": str(ros),
        "hit_rate_0.1pct": str(hit),
        "n_observations": str(n_obs),
    }


def test_rank_top_n_sorts_ascending_by_rmse_over_spread():
    stats = [_stat(11, 0.5), _stat(20, 0.1), _stat(35, 0.3)]
    ranked = rank_top_n(stats, n=10, excluded=set())
    assert [r["publisher_id"] for r in ranked] == ["20", "35", "11"]


def test_rank_top_n_takes_top_n_only():
    stats = [_stat(i, i * 0.01) for i in range(20)]
    ranked = rank_top_n(stats, n=5, excluded=set())
    assert len(ranked) == 5
    assert [r["publisher_id"] for r in ranked] == ["0", "1", "2", "3", "4"]


def test_rank_top_n_excludes_excluded_publishers():
    stats = [_stat(11, 0.1), _stat(23, 0.05), _stat(20, 0.2)]
    ranked = rank_top_n(stats, n=10, excluded={23})
    assert [r["publisher_id"] for r in ranked] == ["11", "20"]


def test_rank_top_n_skips_rows_with_bad_rmse_over_spread(capsys):
    stats = [
        _stat(11, 0.1),
        {
            "publisher_id": "20",
            "rmse_over_spread": "abc",
            "hit_rate_0.1pct": "0",
            "n_observations": "0",
        },
        _stat(35, 0.2),
    ]
    ranked = rank_top_n(stats, n=10, excluded=set())
    assert [r["publisher_id"] for r in ranked] == ["11", "35"]


from lazer_dq.summarize_feeds import apply_filter
from lazer_dq.summarize_feeds import _format_mult, _topup_note


# ---------- _format_mult / _topup_note ----------


def test_format_mult_drops_trailing_zero():
    assert _format_mult(2.0) == "2"
    assert _format_mult(1.5) == "1.5"
    assert _format_mult(3.0) == "3"


def test_topup_note_renders_counts_and_multiplier():
    assert _topup_note(2, 3, 2.0) == "2 passed + 3 top-up (≤2×)"
    assert _topup_note(0, 5, 2.0) == "0 passed + 5 top-up (≤2×)"
    assert _topup_note(1, 4, 1.5) == "1 passed + 4 top-up (≤1.5×)"


# ---------- apply_filter ----------


def test_apply_filter_returns_all_passers_when_at_or_above_floor():
    # 6 publishers all pass; floor is a minimum, never a cap -> return all 6.
    stats = [
        _stat(11, 0.5),
        _stat(20, 0.3),
        _stat(35, 0.4),
        _stat(42, 0.2),
        _stat(50, 0.6),
        _stat(60, 0.1),
    ]
    selected, n_passed, n_topup = apply_filter(
        stats, max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert n_passed == 6
    assert n_topup == 0
    # All returned, sorted ascending by rmse_over_spread.
    assert [r["publisher_id"] for r in selected] == ["60", "42", "20", "35", "11", "50"]


def test_apply_filter_tops_up_to_floor_with_near_misses():
    stats = [
        _stat(11, 0.5),  # passer
        _stat(20, 0.3),  # passer
        _stat(35, 1.4),  # near-miss (r/s > 1.0 but <= 2.0 ceiling)
        _stat(42, 1.8),  # near-miss
        _stat(50, 1.2),  # near-miss
        _stat(60, 1.9),  # near-miss (6th best, not needed once floor reached)
    ]
    selected, n_passed, n_topup = apply_filter(
        stats, max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert n_passed == 2
    assert n_topup == 3
    ids = {r["publisher_id"] for r in selected}
    # passers 11, 20 + 3 best near-misses by r/s: 50 (1.2), 35 (1.4), 42 (1.8).
    assert ids == {"11", "20", "50", "35", "42"}
    assert "60" not in ids


def test_apply_filter_ceiling_excludes_bad_topups_even_below_floor():
    stats = [
        _stat(11, 0.5),  # passer
        _stat(20, 1.5),  # near-miss within 2.0 ceiling
        _stat(35, 2.5),  # over ceiling -> never promoted
        _stat(42, 3.0),  # over ceiling -> never promoted
    ]
    selected, n_passed, n_topup = apply_filter(
        stats, max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert n_passed == 1
    assert n_topup == 1
    # Stays below the floor of 5; 35 and 42 are never promoted.
    assert {r["publisher_id"] for r in selected} == {"11", "20"}


def test_apply_filter_topups_must_meet_n_obs_floor():
    stats = [
        _stat(11, 0.5, n_obs=10000),  # passer
        _stat(20, 1.5, n_obs=500),  # within ceiling but too few observations
        _stat(35, 1.6, n_obs=10000),  # eligible near-miss
    ]
    selected, n_passed, n_topup = apply_filter(
        stats, max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert n_passed == 1
    assert n_topup == 1
    assert {r["publisher_id"] for r in selected} == {"11", "35"}


def test_apply_filter_hit_rate_does_not_gate_topups():
    # All fail hit-rate (10 < 80) -> 0 passers. 11 passes r/s but fails hit,
    # so it is a non-passer that is still eligible as a top-up.
    stats = [
        _stat(11, 0.5, hit=10),
        _stat(20, 1.1, hit=10),
        _stat(35, 1.5, hit=10),
        _stat(42, 1.9, hit=10),
        _stat(50, 1.3, hit=10),
        _stat(60, 1.4, hit=10),
    ]
    selected, n_passed, n_topup = apply_filter(
        stats, max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert n_passed == 0
    assert n_topup == 5
    # Best 5 by r/s: 11 (0.5), 20 (1.1), 50 (1.3), 60 (1.4), 35 (1.5); 42 (1.9) excluded.
    assert {r["publisher_id"] for r in selected} == {"11", "20", "50", "60", "35"}
    assert "42" not in {r["publisher_id"] for r in selected}


def test_apply_filter_empty_when_all_over_ceiling():
    stats = [
        _stat(11, 2.5, hit=10),
        _stat(20, 3.0, hit=10),
        _stat(35, 5.0, hit=10),
    ]
    selected, n_passed, n_topup = apply_filter(
        stats, max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert selected == []
    assert n_passed == 0
    assert n_topup == 0


def test_apply_filter_returns_empty_for_empty_input():
    selected, n_passed, n_topup = apply_filter(
        [], max_ros=1.0, min_hit=80, min_obs=1000, floor=5, ceiling_mult=2.0
    )
    assert selected == []
    assert n_passed == 0
    assert n_topup == 0


from lazer_dq.summarize_feeds import compute_aggregate


# ---------- compute_aggregate ----------


def test_compute_aggregate_is_sorted_union_of_per_session_arrays():
    arrays = [[11, 20, 35], [20, 22, 41], [11, 42]]
    assert compute_aggregate(arrays) == [11, 20, 22, 35, 41, 42]


def test_compute_aggregate_skips_none_sessions():
    arrays = [[11, 20], None, [22, 11]]
    assert compute_aggregate(arrays) == [11, 20, 22]


def test_compute_aggregate_empty_when_all_sessions_empty():
    assert compute_aggregate([None, None, None, None]) == []
    assert compute_aggregate([[], [], []]) == []


def test_compute_aggregate_deduplicates():
    arrays = [[11, 11, 20], [20, 20]]
    assert compute_aggregate(arrays) == [11, 20]


# ---------- main (integration) ----------

import sys

from openpyxl import load_workbook

from lazer_dq.summarize_feeds import main


def test_main_writes_workbook_for_one_feed_one_mode(tmp_path, monkeypatch, capsys):
    """End-to-end happy path: 1 feed, 1 mode populated, 3 missing modes."""
    # publishers.md
    pubs_md = tmp_path / "publishers.md"
    pubs_md.write_text(
        """\
| ID  | Name              | Active |
| --- | ----------------- | ------ |
| 11  | Amber.Production  | Yes    |
| 23  | LoTech.Test       | Yes    |
"""
    )

    # CSV with one feed.
    csv = tmp_path / "input.csv"
    csv.write_text("1021, 2026-05-06, us-equities-post\n")

    # dq_reports tree - only us-equities-post for feed 1021 has data.
    reports = tmp_path / "dq_reports"
    _write_stats_csv(
        reports,
        "lazer-prod",
        "us-equities-post",
        1021,
        "2026-05-06",
        [
            "1021,11,22218,-0.05,0.08,-0.01,0.02,0.0932,0.51,0.0185,0.07,-84,0,75,0,0,100.0,0.96,fail\n",
            # excluded .Test publisher 23 - must not appear anywhere.
            "1021,23,5000,-0.05,0.08,-0.01,0.02,0.05,0.5,0.01,0.05,0,0,0,0,0,100.0,0.5,fail\n",
        ],
    )

    out_path = tmp_path / "out.xlsx"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-06",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(pubs_md),
            "--output",
            str(out_path),
        ],
    )

    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 0

    # Workbook exists with both sheets.
    wb = load_workbook(out_path, data_only=True)
    assert "rankings" in wb.sheetnames
    assert "allowed" in wb.sheetnames

    # Allowed sheet: 1 aggregate row + 4 session rows = 5 data rows starting at row 3.
    allow = wb["allowed"]
    assert allow.cell(3, 1).value == 1021
    assert allow.cell(3, 2).value == "(aggregate)"
    # Aggregate JSON contains publisher 11 only (23 excluded as .Test).
    assert allow.cell(3, 3).value == '"allowedPublisherIds": [ 11 ],'
    # Session rows in MODE_ORDER: us-equities (no data), pre (no data), post (data), overnight (no data).
    assert allow.cell(4, 2).value == "REGULAR"
    assert allow.cell(4, 3).value == "(no data)"
    assert allow.cell(5, 2).value == "PRE_MARKET"
    assert allow.cell(6, 2).value == "POST_MARKET"
    assert allow.cell(6, 3).value == '"allowedPublisherIds": [ 11 ],'
    assert allow.cell(7, 2).value == "OVER_NIGHT"
    assert allow.cell(7, 3).value == "(no data)"

    # Rankings sheet: feed banner + at least 1 data row.
    rank = wb["rankings"]
    found_banner = any(
        rank.cell(r, 1).value == "=== Feed 1021 ===" for r in range(1, 30)
    )
    assert found_banner


def test_main_skipped_feeds_section_lists_zero_data_feeds(
    tmp_path, monkeypatch, capsys
):
    """Feed in CSV with no data anywhere → listed in skipped footer + stdout summary."""
    pubs_md = tmp_path / "publishers.md"
    pubs_md.write_text(
        "| ID | Name | Active |\n| --- | --- | --- |\n| 11 | Amber.Production | Yes |\n"
    )

    csv = tmp_path / "input.csv"
    csv.write_text("1021, 2026-05-06, us-equities\n9999, 2026-05-06, us-equities\n")

    reports = tmp_path / "dq_reports"
    _write_stats_csv(
        reports,
        "lazer-prod",
        "us-equities",
        1021,
        "2026-05-06",
        [
            "1021,11,22218,-0.05,0.08,-0.01,0.02,0.0932,0.51,0.0185,0.07,-84,0,75,0,0,100.0,0.96,fail\n"
        ],
    )
    # Feed 9999 has no stats anywhere.

    out_path = tmp_path / "out.xlsx"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-06",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(pubs_md),
            "--output",
            str(out_path),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 0

    out = capsys.readouterr().out
    assert "9999" in out
    assert "Feeds skipped" in out

    wb = load_workbook(out_path, data_only=True)
    allow = wb["allowed"]
    # Find footer header.
    found_footer_label = False
    found_9999 = False
    for r in range(1, 60):
        v = allow.cell(r, 1).value
        if v == "Feeds skipped (no data for any mode):":
            found_footer_label = True
        if v == 9999:
            found_9999 = True
    assert found_footer_label
    assert found_9999


def test_main_no_data_anywhere_exits_nonzero(tmp_path, monkeypatch):
    pubs_md = tmp_path / "publishers.md"
    pubs_md.write_text("| ID | Name | Active |\n| --- | --- | --- |\n")

    csv = tmp_path / "input.csv"
    csv.write_text("1021, 2026-05-06, us-equities\n")

    reports = tmp_path / "dq_reports"
    (reports / "lazer-prod").mkdir(parents=True)  # cluster dir exists but empty

    out_path = tmp_path / "out.xlsx"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-06",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(pubs_md),
            "--output",
            str(out_path),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 1


def test_main_excluded_publishers_never_appear_in_either_sheet(tmp_path, monkeypatch):
    """A .Test publisher with stellar metrics must not appear in rankings or allowed."""
    pubs_md = tmp_path / "publishers.md"
    pubs_md.write_text(
        """\
| ID | Name             | Active |
| --- | --------------- | ------ |
| 11 | Amber.Production | Yes   |
| 23 | LoTech.Test      | Yes   |
"""
    )

    csv = tmp_path / "input.csv"
    csv.write_text("1021, 2026-05-06, us-equities\n")

    reports = tmp_path / "dq_reports"
    _write_stats_csv(
        reports,
        "lazer-prod",
        "us-equities",
        1021,
        "2026-05-06",
        [
            # publisher 23 has the BEST rmse_over_spread but is .Test → must be filtered out.
            "1021,23,99999,-0.001,0.001,-0.0001,0.0001,0.001,0.001,0.0001,0.001,0,0,0,0,0,100.0,0.01,fail\n",
            "1021,11,22218,-0.05,0.08,-0.01,0.02,0.0932,0.51,0.0185,0.07,-84,0,75,0,0,100.0,0.96,fail\n",
        ],
    )

    out_path = tmp_path / "out.xlsx"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-06",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(pubs_md),
            "--output",
            str(out_path),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 0

    wb = load_workbook(out_path, data_only=True)

    # Rankings sheet: scan all cells for "23" — none should appear as a publisher_id.
    rank = wb["rankings"]
    for r in range(1, 30):
        for c in range(1, 25):
            assert (
                rank.cell(r, c).value != 23
            ), f"excluded publisher 23 leaked into rankings at ({r},{c})"

    # Allowed sheet: column C JSON arrays must not contain 23.
    import re as _re

    allow = wb["allowed"]
    for r in range(1, 30):
        v = allow.cell(r, 3).value
        if not isinstance(v, str) or "allowedPublisherIds" not in v:
            continue
        m = _re.search(r"\[(.*?)\]", v)
        if not m or not m.group(1).strip():
            continue
        ids = [int(x.strip()) for x in m.group(1).split(",")]
        assert 23 not in ids, f"excluded publisher 23 leaked into allowed JSON: {v}"


def test_main_missing_csv_exits_nonzero(tmp_path, monkeypatch, capsys):
    pubs_md = tmp_path / "publishers.md"
    pubs_md.write_text("| ID | Name | Active |\n| --- | --- | --- |\n")

    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(tmp_path / "missing.csv"),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-06",
            "--publishers-md",
            str(pubs_md),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 1
    assert "not found" in capsys.readouterr().out


# ---------- ASSET_CLASS_CONFIG registry ----------

from lazer_dq.summarize_feeds import ASSET_CLASS_CONFIG


def test_registry_has_us_equities_entry_with_all_required_keys():
    assert "us-equities" in ASSET_CLASS_CONFIG
    cfg = ASSET_CLASS_CONFIG["us-equities"]
    assert cfg["modes"] == [
        "us-equities",
        "us-equities-pre",
        "us-equities-post",
        "us-equities-overnight",
    ]
    assert cfg["sessions"] == {
        "us-equities": "REGULAR",
        "us-equities-pre": "PRE_MARKET",
        "us-equities-post": "POST_MARKET",
        "us-equities-overnight": "OVER_NIGHT",
    }
    assert cfg["default_max_ros"] == {
        "us-equities": 1.0,
        "us-equities-pre": 2.0,
        "us-equities-post": 2.0,
        "us-equities-overnight": 3.0,
    }
    assert cfg["default_min_hit"] == {
        "us-equities": 80.0,
        "us-equities-pre": 50.0,
        "us-equities-post": 50.0,
        "us-equities-overnight": 25.0,
    }


def test_registry_has_hk_equities_entry():
    assert "hk-equities" in ASSET_CLASS_CONFIG
    cfg = ASSET_CLASS_CONFIG["hk-equities"]
    assert cfg["modes"] == ["hk-equities"]
    assert cfg["sessions"] == {"hk-equities": "REGULAR"}
    assert cfg["default_max_ros"] == {"hk-equities": 1.0}
    assert cfg["default_min_hit"] == {"hk-equities": 80.0}


def test_legacy_constants_still_match_us_equities_registry_entry():
    """Back-compat: MODE_ORDER / MODE_TO_SESSION still exist for any external importer."""
    from lazer_dq.summarize_feeds import MODE_ORDER, MODE_TO_SESSION

    assert MODE_ORDER == ASSET_CLASS_CONFIG["us-equities"]["modes"]
    assert MODE_TO_SESSION == ASSET_CLASS_CONFIG["us-equities"]["sessions"]


# ---------- _build_per_feed_data with custom modes ----------

from lazer_dq.summarize_feeds import _build_per_feed_data


def test_build_per_feed_data_honors_modes_parameter(tmp_path):
    """Only the modes passed in are looked up under reports_dir; others are not touched."""
    reports = tmp_path / "dq_reports"
    # Write a stats.csv ONLY for hk-equities.
    _write_stats_csv(
        reports,
        "lazer-prod",
        "hk-equities",
        884,
        "2026-05-19",
        body_rows=["5,5000,0.001,0.5,90.0\n"],
        header="publisher_id,n_observations,rmse,rmse_over_spread,hit_rate_0.1pct\n",
    )
    (
        per_feed,
        skipped,
        topup_rows,
        zero_passer_rows,
        modes_with_data,
    ) = _build_per_feed_data(
        feed_ids=[884],
        reports_dir=reports,
        cluster="lazer-prod",
        date="2026-05-19",
        excluded={0},
        top_n=10,
        max_ros_map={"hk-equities": 1.0},
        min_hit_map={"hk-equities": 80.0},
        min_obs=1000,
        floor=5,
        ceiling_mult=2.0,
        modes=["hk-equities"],
    )
    assert skipped == []
    assert modes_with_data == 1
    assert per_feed[884]["hk-equities"] is not None
    assert per_feed[884]["hk-equities"]["ranked"][0]["publisher_id"] == "5"
    # Publisher 5 (r/s 0.5, hit 90, 5000 obs) passes outright.
    assert per_feed[884]["hk-equities"]["n_passed"] == 1
    assert per_feed[884]["hk-equities"]["n_topup"] == 0
    assert topup_rows == 0
    assert zero_passer_rows == 0
    # Crucially: no us-equities key at all.
    assert "us-equities" not in per_feed[884]


# ---------- write_rankings_sheet parametric layout ----------

from lazer_dq.summarize_feeds import write_rankings_sheet


def _ranked_row(pub_id, n_obs=5000, rmse=0.001, ros=0.5, hit=90.0):
    return {
        "publisher_id": str(pub_id),
        "n_observations": str(n_obs),
        "rmse": str(rmse),
        "rmse_over_spread": str(ros),
        "hit_rate_0.1pct": str(hit),
    }


def test_write_rankings_sheet_one_mode_uses_6_columns(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    per_feed = {
        884: {
            "hk-equities": {
                "ranked": [_ranked_row(5), _ranked_row(7)],
                "filtered": [_ranked_row(5)],
                "n_passed": 2,
                "n_topup": 0,
            }
        }
    }
    write_rankings_sheet(
        ws,
        per_feed,
        date="2026-05-19",
        cluster="lazer-prod",
        modes=["hk-equities"],
    )
    out = tmp_path / "out.xlsx"
    wb.save(out)

    from openpyxl import load_workbook

    wb2 = load_workbook(out)
    ws2 = wb2["Sheet"]
    assert ws2.cell(row=3, column=2).value == "hk-equities"
    assert ws2.cell(row=4, column=1).value == "rank"
    assert ws2.cell(row=4, column=2).value == "pub"
    assert ws2.cell(row=4, column=6).value == "hit%"
    assert ws2.cell(row=4, column=7).value is None
    assert ws2.cell(row=7, column=1).value == 1
    assert ws2.cell(row=7, column=2).value == 5


def test_write_rankings_sheet_four_modes_uses_24_columns(tmp_path):
    """Regression: us-equities layout is unchanged (24 cols, 5-col blocks + spacers at G/M/S)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    per_feed = {
        922: {
            "us-equities": {
                "ranked": [_ranked_row(5)],
                "filtered": [_ranked_row(5)],
                "n_passed": 2,
                "n_topup": 0,
            },
            "us-equities-pre": None,
            "us-equities-post": None,
            "us-equities-overnight": None,
        }
    }
    write_rankings_sheet(
        ws,
        per_feed,
        date="2026-05-19",
        cluster="lazer-prod",
        modes=[
            "us-equities",
            "us-equities-pre",
            "us-equities-post",
            "us-equities-overnight",
        ],
    )
    assert ws.cell(row=3, column=2).value == "us-equities"
    assert ws.cell(row=3, column=8).value == "us-equities-pre"
    assert ws.cell(row=3, column=14).value == "us-equities-post"
    assert ws.cell(row=3, column=20).value == "us-equities-overnight"


# ---------- write_allowed_sheet parametric layout ----------

from lazer_dq.summarize_feeds import write_allowed_sheet


def test_write_allowed_sheet_one_mode_emits_two_rows_per_feed(tmp_path):
    """For hk-equities (1 mode): each feed gets 1 aggregate + 1 session row."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    per_feed = {
        884: {
            "hk-equities": {
                "ranked": [_ranked_row(5), _ranked_row(7)],
                "filtered": [_ranked_row(5), _ranked_row(7)],
                "n_passed": 2,
                "n_topup": 0,
            }
        }
    }
    write_allowed_sheet(
        ws,
        per_feed,
        skipped_feeds=[],
        date="2026-05-19",
        cluster="lazer-prod",
        modes=["hk-equities"],
        sessions={"hk-equities": "REGULAR"},
        ceiling_mult=2.0,
    )
    assert ws.cell(row=3, column=1).value == 884
    assert ws.cell(row=3, column=2).value == "(aggregate)"
    assert "5, 7" in ws.cell(row=3, column=3).value
    assert ws.cell(row=4, column=1).value == 884
    assert ws.cell(row=4, column=2).value == "REGULAR"
    assert "5, 7" in ws.cell(row=4, column=3).value


def test_write_allowed_sheet_topup_note_when_below_floor(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    per_feed = {
        884: {
            "hk-equities": {
                "ranked": [_ranked_row(5)],
                "filtered": [_ranked_row(5), _ranked_row(7), _ranked_row(9)],
                "n_passed": 1,
                "n_topup": 2,
            }
        }
    }
    write_allowed_sheet(
        ws,
        per_feed,
        skipped_feeds=[],
        date="2026-05-19",
        cluster="lazer-prod",
        modes=["hk-equities"],
        sessions={"hk-equities": "REGULAR"},
        ceiling_mult=2.0,
    )
    # Session row is row 4; Notes is column 4.
    assert ws.cell(row=4, column=4).value == "1 passed + 2 top-up (≤2×)"
    assert "5, 7, 9" in ws.cell(row=4, column=3).value


def test_write_allowed_sheet_all_passers_has_blank_note(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    per_feed = {
        884: {
            "hk-equities": {
                "ranked": [_ranked_row(5)],
                "filtered": [_ranked_row(5), _ranked_row(7)],
                "n_passed": 2,
                "n_topup": 0,
            }
        }
    }
    write_allowed_sheet(
        ws,
        per_feed,
        skipped_feeds=[],
        date="2026-05-19",
        cluster="lazer-prod",
        modes=["hk-equities"],
        sessions={"hk-equities": "REGULAR"},
        ceiling_mult=2.0,
    )
    assert ws.cell(row=4, column=4).value is None


def test_write_allowed_sheet_empty_by_ceiling_note(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    per_feed = {
        884: {
            "hk-equities": {
                "ranked": [_ranked_row(5, ros=3.0)],
                "filtered": [],
                "n_passed": 0,
                "n_topup": 0,
            }
        }
    }
    write_allowed_sheet(
        ws,
        per_feed,
        skipped_feeds=[],
        date="2026-05-19",
        cluster="lazer-prod",
        modes=["hk-equities"],
        sessions={"hk-equities": "REGULAR"},
        ceiling_mult=2.0,
    )
    assert ws.cell(row=4, column=3).value == "(no data)"
    assert ws.cell(row=4, column=4).value == "0 passed, all > 2× ceiling"


# ---------- validate_csv_modes ----------

from lazer_dq.summarize_feeds import validate_csv_modes


def test_validate_csv_modes_accepts_matching_modes(tmp_path):
    csv = tmp_path / "ok.csv"
    csv.write_text("884,2026-05-19,hk-equities\n" "885,2026-05-19,hk-equities\n")
    assert validate_csv_modes(csv, allowed_modes=["hk-equities"]) is None


def test_validate_csv_modes_accepts_empty_third_column(tmp_path):
    """Back-compat: feed-id-only CSVs (no mode column) are still accepted."""
    csv = tmp_path / "legacy.csv"
    csv.write_text("884\n885\n")
    assert validate_csv_modes(csv, allowed_modes=["hk-equities"]) is None


def test_validate_csv_modes_rejects_mismatched_modes(tmp_path, capsys):
    csv = tmp_path / "bad.csv"
    csv.write_text("884,2026-05-19,us-equities\n" "885,2026-05-19,us-equities\n")
    with pytest.raises(SystemExit) as exc:
        validate_csv_modes(csv, allowed_modes=["hk-equities"])
    assert exc.value.code != 0
    out = capsys.readouterr().out
    assert "us-equities" in out
    assert "hk-equities" in out


# ---------- main() with --asset-class hk-equities ----------


def test_main_hk_equities_end_to_end(tmp_path, monkeypatch):
    """Full run for a 1-feed hk-equities CSV produces a workbook with the HK layout."""
    reports = tmp_path / "dq_reports"
    _write_stats_csv(
        reports,
        "lazer-prod",
        "hk-equities",
        884,
        "2026-05-19",
        body_rows=[
            "5,5000,0.001,0.5,90.0\n",
            "7,5000,0.001,0.6,92.0\n",
        ],
        header="publisher_id,n_observations,rmse,rmse_over_spread,hit_rate_0.1pct\n",
    )
    csv = tmp_path / "hk.csv"
    csv.write_text("884,2026-05-19,hk-equities\n")
    md = tmp_path / "publishers.md"
    md.write_text("| ID | Name | Active |\n| 0 | Zero.Test | Yes |\n")
    out = tmp_path / "out.xlsx"

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-19",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(md),
            "--asset-class",
            "hk-equities",
            "--output",
            str(out),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 0
    assert out.exists()

    wb = load_workbook(out)
    rank = wb["rankings"]
    assert rank.cell(row=3, column=2).value == "hk-equities"
    assert rank.cell(row=4, column=7).value is None

    allowed = wb["allowed"]
    assert allowed.cell(row=3, column=2).value == "(aggregate)"
    assert allowed.cell(row=4, column=2).value == "REGULAR"


def test_main_rejects_mode_mismatch(tmp_path, monkeypatch, capsys):
    """--asset-class hk-equities + CSV containing us-equities rows -> exit non-zero."""
    csv = tmp_path / "mixed.csv"
    csv.write_text("884,2026-05-19,us-equities\n")
    md = tmp_path / "publishers.md"
    md.write_text("# empty\n")
    reports = tmp_path / "dq_reports"
    (reports / "lazer-prod").mkdir(parents=True)

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-19",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(md),
            "--asset-class",
            "hk-equities",
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code != 0
    out = capsys.readouterr().out
    assert "us-equities" in out


def test_main_hk_equities_topup_path_end_to_end(tmp_path, monkeypatch, capsys):
    """Full run: 1 passer + near-misses -> floor tops up. Verifies the yellow
    Notes cell, the ceiling exclusion, and the new stdout counters end to end."""
    reports = tmp_path / "dq_reports"
    _write_stats_csv(
        reports,
        "lazer-prod",
        "hk-equities",
        884,
        "2026-05-19",
        body_rows=[
            "5,5000,0.001,0.5,90.0\n",  # passer (r/s 0.5 <= 1.0)
            "7,5000,0.002,1.3,90.0\n",  # top-up (1.0 < r/s <= 2.0 ceiling)
            "9,5000,0.002,1.7,90.0\n",  # top-up
            "11,5000,0.002,1.9,90.0\n",  # top-up
            "13,5000,0.002,1.95,90.0\n",  # top-up (reaches floor of 5)
            "15,5000,0.002,2.5,90.0\n",  # over 2x ceiling -> never promoted
        ],
        header="publisher_id,n_observations,rmse,rmse_over_spread,hit_rate_0.1pct\n",
    )
    csv = tmp_path / "hk.csv"
    csv.write_text("884,2026-05-19,hk-equities\n")
    md = tmp_path / "publishers.md"
    md.write_text("| ID | Name | Active |\n| 0 | Zero.Test | Yes |\n")
    out = tmp_path / "out.xlsx"

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "summarize_feeds",
            "--csv",
            str(csv),
            "--cluster",
            "lazer-prod",
            "--date",
            "2026-05-19",
            "--reports-dir",
            str(reports),
            "--publishers-md",
            str(md),
            "--asset-class",
            "hk-equities",
            "--output",
            str(out),
        ],
    )
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 0

    captured = capsys.readouterr().out
    assert "Rows using top-ups: 1 cells" in captured
    assert "Rows with 0 passers: 0 cells" in captured

    wb = load_workbook(out, data_only=True)
    allowed = wb["allowed"]
    # Row 3 = aggregate, row 4 = REGULAR session (hk-equities has one mode).
    assert allowed.cell(row=4, column=2).value == "REGULAR"
    json_cell = allowed.cell(row=4, column=3).value
    assert "5, 7, 9, 11, 13" in json_cell
    assert "15" not in json_cell  # over the 2x ceiling, never promoted
    assert allowed.cell(row=4, column=4).value == "1 passed + 4 top-up (≤2×)"
