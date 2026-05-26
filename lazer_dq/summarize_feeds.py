#!/usr/bin/env python3
"""DQ summary workbook generator — reads dq_reports/, emits one .xlsx.

Two sheets per run:
  rankings — top-N publishers per (feed, mode) by rmse_over_spread, modes side-by-side
  allowed  — paste-ready allowedPublisherIds JSON arrays per feed/session

Per-asset-class layout. Pick the asset class with --asset-class:
  us-equities (default) — 4 modes: regular, pre, post, overnight (24-col layout)
  hk-equities           — 1 mode:  regular (6-col layout)

Adding a new asset class = adding one entry to ASSET_CLASS_CONFIG.

Run:
    python3 -m lazer_dq.summarize_feeds \
        --csv MV_Mario_3_pre.csv --cluster lazer-prod --date 2026-05-06

    python3 -m lazer_dq.summarize_feeds \
        --csv equity_hk_feed_ids.csv --asset-class hk-equities \
        --cluster lazer-prod --date 2026-05-19
"""
import argparse
import csv as csv_mod
import sys
from pathlib import Path

# Asset-class registry. Adding a new asset class = adding one entry here.
# Each entry declares:
#   modes:           ordered list of dq_reports/<cluster>/<mode>/ directory names to read.
#   sessions:        mode -> after.json session-label, for the 'allowed' sheet.
#   default_max_ros: per-mode max rmse_over_spread threshold.
#   default_min_hit: per-mode min hit_rate_0.1pct (%) threshold.
ASSET_CLASS_CONFIG: dict = {
    "us-equities": {
        "modes": [
            "us-equities",
            "us-equities-pre",
            "us-equities-post",
            "us-equities-overnight",
        ],
        "sessions": {
            "us-equities": "REGULAR",
            "us-equities-pre": "PRE_MARKET",
            "us-equities-post": "POST_MARKET",
            "us-equities-overnight": "OVER_NIGHT",
        },
        "default_max_ros": {
            "us-equities": 1.0,
            "us-equities-pre": 2.0,
            "us-equities-post": 2.0,
            "us-equities-overnight": 3.0,
        },
        "default_min_hit": {
            "us-equities": 80.0,
            "us-equities-pre": 50.0,
            "us-equities-post": 50.0,
            "us-equities-overnight": 25.0,
        },
    },
    "hk-equities": {
        "modes": ["hk-equities"],
        "sessions": {"hk-equities": "REGULAR"},
        "default_max_ros": {"hk-equities": 1.0},
        "default_min_hit": {"hk-equities": 80.0},
    },
}

# Back-compat aliases — kept so any external code importing these names keeps working.
# Internal code should prefer ASSET_CLASS_CONFIG[<slug>][...] going forward.
MODE_TO_SESSION = ASSET_CLASS_CONFIG["us-equities"]["sessions"]
MODE_ORDER = ASSET_CLASS_CONFIG["us-equities"]["modes"]

DEFAULT_MIN_N_OBS = 1000
DEFAULT_TOP_N = 10
DEFAULT_REDUNDANCY_FLOOR = 5
DEFAULT_TOPUP_CEILING_MULT = 2.0


def load_excluded_publishers(publishers_md_path) -> set[int]:
    """Parse publishers.md markdown table, return IDs to exclude.

    Excluded = {0} ∪ {ids whose Name ends with ".Test"}.
    Malformed rows are skipped silently. Always includes 0 even if file is empty.
    """
    excluded: set[int] = {0}
    with open(publishers_md_path, "r") as f:
        for line in f:
            line = line.strip()
            if not line.startswith("|"):
                continue
            # Strip leading/trailing pipes, split on | and trim each cell.
            parts = [p.strip() for p in line.strip("|").split("|")]
            if len(parts) < 2:
                continue
            try:
                pub_id = int(parts[0])
            except ValueError:
                # Header row ("ID"), separator row ("---"), or malformed; skip.
                continue
            name = parts[1]
            if name.endswith(".Test"):
                excluded.add(pub_id)
    return excluded


def discover_feeds(csv_path) -> list[int]:
    """Distinct numeric feed_ids from CSV column 1, in first-seen order.

    Empty rows, rows with empty first column, and rows with non-numeric
    feed_ids are skipped with a stdout warning.
    """
    seen: list[int] = []
    seen_set: set[int] = set()
    with open(csv_path, "r") as f:
        reader = csv_mod.reader(f)
        for row in reader:
            if not row or not row[0].strip():
                continue
            raw = row[0].strip()
            try:
                feed_id = int(raw)
            except ValueError:
                print(
                    f"  Warning: skipping malformed CSV row (non-numeric feed_id): {raw!r}"
                )
                continue
            if feed_id not in seen_set:
                seen.append(feed_id)
                seen_set.add(feed_id)
    return seen


def validate_csv_modes(csv_path, allowed_modes: list) -> None:
    """Verify every CSV row's column-3 mode is in allowed_modes.

    Rows with empty column 3 (legacy feed-id-only CSVs) are accepted.
    On mismatch, print an explanatory error and sys.exit(1).
    """
    bad: list = []
    allowed_set = set(allowed_modes)
    with open(csv_path, "r") as f:
        reader = csv_mod.reader(f)
        for row in reader:
            if not row or not row[0].strip():
                continue
            if len(row) < 3:
                continue  # legacy feed-id-only row
            mode = row[2].strip()
            if not mode:
                continue
            if mode not in allowed_set:
                bad.append((row[0].strip(), mode))
    if bad:
        sample = ", ".join(f"{fid} ({m})" for fid, m in bad[:5])
        more = f" (and {len(bad) - 5} more)" if len(bad) > 5 else ""
        print(
            f"Error: CSV contains modes not in --asset-class={allowed_modes!r}.\n"
            f"       Allowed modes: {sorted(allowed_set)}\n"
            f"       Mismatched rows: {sample}{more}"
        )
        sys.exit(1)


def load_stats(reports_dir, cluster: str, mode: str, feed_id: int, date: str):
    """Read dq_reports/<cluster>/<mode>/<feed_id>/<date>/stats.csv.

    Returns a list of dicts (csv.DictReader output), or None if the file is missing.
    """
    path = Path(reports_dir) / cluster / mode / str(feed_id) / date / "stats.csv"
    if not path.exists():
        return None
    with open(path, "r") as f:
        return list(csv_mod.DictReader(f))


def rank_top_n(stats, n: int, excluded: set[int]) -> list[dict]:
    """Drop excluded publisher_ids, sort ascending by rmse_over_spread, take top n.

    Rows with non-numeric publisher_id or rmse_over_spread are skipped with a warning.
    """
    keyed: list[tuple[float, dict]] = []
    for r in stats:
        try:
            pid = int(r["publisher_id"])
        except (ValueError, KeyError):
            continue
        if pid in excluded:
            continue
        try:
            ros = float(r["rmse_over_spread"])
        except (ValueError, KeyError):
            print(
                f"  Warning: skipping row with bad rmse_over_spread: publisher_id={r.get('publisher_id')}"
            )
            continue
        keyed.append((ros, r))
    keyed.sort(key=lambda x: x[0])
    return [r for _, r in keyed[:n]]


def apply_filter(
    stats, max_ros: float, min_hit: float, min_obs: int, floor: int, ceiling_mult: float
):
    """Apply per-mode thresholds with a redundancy floor. Return (selected, n_passed, n_topup).

    selected : passers (sorted ascending by rmse_over_spread) plus, when there
               are fewer than `floor` passers, the next-best below-threshold
               publishers ("top-ups") sorted by rmse_over_spread. Each top-up
               must clear the n_observations floor AND have
               rmse_over_spread <= ceiling_mult * max_ros. The floor is a
               minimum, never a cap: if more than `floor` publishers pass, all
               of them are returned.
    n_passed : count meeting all three thresholds (r/s, hit_rate, n_obs).
    n_topup  : count of below-threshold fillers added to reach the floor.

    Empty input -> ([], 0, 0). Rows with non-numeric metric fields are skipped.
    Note: hit_rate gates passers only, not top-ups; the ceiling is the top-up
    quality proxy.
    """
    if not stats:
        return [], 0, 0

    passers: list[tuple[float, dict]] = []
    non_passers: list[tuple[float, dict, int]] = []
    for r in stats:
        try:
            ros = float(r["rmse_over_spread"])
            hit = float(r["hit_rate_0.1pct"])
            n_obs = int(r["n_observations"])
        except (ValueError, KeyError):
            continue
        if ros <= max_ros and hit >= min_hit and n_obs >= min_obs:
            passers.append((ros, r))
        else:
            non_passers.append((ros, r, n_obs))

    passers.sort(key=lambda x: x[0])
    n_passed = len(passers)
    if n_passed >= floor:
        return [r for _, r in passers], n_passed, 0

    # Top up with below-threshold publishers within the quality ceiling.
    ceiling = ceiling_mult * max_ros
    eligible = [
        (ros, r)
        for (ros, r, n_obs) in non_passers
        if n_obs >= min_obs and ros <= ceiling
    ]
    eligible.sort(key=lambda x: x[0])
    topups = eligible[: floor - n_passed]
    selected = [r for _, r in passers] + [r for _, r in topups]
    return selected, n_passed, len(topups)


def compute_aggregate(per_session_arrays) -> list[int]:
    """Sorted union of per-session publisher_id arrays.

    None entries (mode missing) are skipped. Empty list if every session is empty/None.
    """
    union: set[int] = set()
    for arr in per_session_arrays:
        if arr is None:
            continue
        union.update(arr)
    return sorted(union)


def write_rankings_sheet(
    ws, per_feed_data: dict, date: str, cluster: str, modes: list
) -> None:
    """Populate the 'rankings' worksheet.

    Layout is parametric on len(modes):
      - 1 rank column (A) +
      - N × 5-col mode blocks (pub | n_obs | rmse | r/s | hit%) +
      - (N-1) × 1-col spacers between blocks
      = 6N total columns.

    Examples:
      - 4 modes (us-equities) → 24 cols, blocks at B/H/N/T (unchanged from prior layout).
      - 1 mode  (hk-equities) → 6 cols, single block at B.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    bold_lg = Font(bold=True, size=12)
    bold_xl = Font(bold=True, size=14)
    gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center = Alignment(horizontal="center")

    n_modes = len(modes)
    total_cols = 6 * n_modes  # 1 rank + 5N blocks + (N-1) spacers = 6N
    # mode_starts[m] is the 1-indexed start column of each 5-col mode block.
    # Block 0 starts at column 2 (B). Each subsequent block starts 6 columns later
    # (5 data cols + 1 spacer).
    mode_starts = {mode: 2 + 6 * i for i, mode in enumerate(modes)}
    sub_headers = ["pub", "n_obs", "rmse", "r/s", "hit%"]

    # Row 1: title.
    ws.cell(row=1, column=1, value=f"DQ Summary — {cluster} — {date}").font = bold_xl
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1).alignment = center

    # Row 3: mode-block headers.
    for mode in modes:
        col = mode_starts[mode]
        c = ws.cell(row=3, column=col, value=mode)
        c.font = bold
        c.fill = gray
        c.alignment = center
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 4)

    # Row 4: sub-headers.
    a4 = ws.cell(row=4, column=1, value="rank")
    a4.font = bold
    a4.fill = gray
    for mode in modes:
        start = mode_starts[mode]
        for i, label in enumerate(sub_headers):
            c = ws.cell(row=4, column=start + i, value=label)
            c.font = bold
            c.fill = gray

    # Freeze header rows.
    ws.freeze_panes = "A5"

    # Per-feed sections.
    row = 6
    for feed_id, mode_data in per_feed_data.items():
        # Banner. Force inline-string type so leading '=' is not parsed as a formula.
        banner = ws.cell(row=row, column=1, value=f"=== Feed {feed_id} ===")
        banner.data_type = "s"
        banner.font = bold_lg
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=total_cols
        )
        row += 1

        ranked_per_mode = {
            m: (mode_data[m]["ranked"] if mode_data.get(m) else None) for m in modes
        }
        n_rows = max((len(r) for r in ranked_per_mode.values() if r), default=0)
        if n_rows == 0:
            ws.cell(row=row, column=2, value="(no data)")
            row += 2
            continue

        for i in range(n_rows):
            ws.cell(row=row + i, column=1, value=i + 1)
        for mode in modes:
            start = mode_starts[mode]
            ranked = ranked_per_mode[mode]
            if ranked is None:
                ws.cell(row=row, column=start, value="(no data)")
                continue
            for i, r in enumerate(ranked):
                ws.cell(row=row + i, column=start + 0, value=int(r["publisher_id"]))
                ws.cell(row=row + i, column=start + 1, value=int(r["n_observations"]))
                ws.cell(row=row + i, column=start + 2, value=round(float(r["rmse"]), 4))
                ws.cell(
                    row=row + i,
                    column=start + 3,
                    value=round(float(r["rmse_over_spread"]), 4),
                )
                ws.cell(
                    row=row + i,
                    column=start + 4,
                    value=round(float(r["hit_rate_0.1pct"]), 2),
                )
        row += n_rows + 1

    # Column widths.
    for col_idx in range(1, total_cols + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 9
    ws.column_dimensions["A"].width = 6  # rank


def _format_mult(ceiling_mult: float) -> str:
    """Render the ceiling multiplier without a trailing .0 (2.0 -> '2', 1.5 -> '1.5')."""
    return f"{ceiling_mult:g}"


def _topup_note(n_passed: int, n_topup: int, ceiling_mult: float) -> str:
    """Notes-column text for a row that needed below-threshold top-ups."""
    return f"{n_passed} passed + {n_topup} top-up (≤{_format_mult(ceiling_mult)}×)"


def _format_allowed_pub_ids(ids: list[int]) -> str:
    """Render `"allowedPublisherIds": [ 1, 2, 3 ],` matching after.json's inline style.

    Trailing comma is included so the snippet pastes cleanly inside
    marketSchedules[] entries.
    """
    return f'"allowedPublisherIds": [ {", ".join(str(i) for i in ids)} ],'


def write_allowed_sheet(
    ws,
    per_feed_data: dict,
    skipped_feeds: list,
    date: str,
    cluster: str,
    modes: list,
    sessions: dict,
    ceiling_mult: float = DEFAULT_TOPUP_CEILING_MULT,
) -> None:
    """Populate the 'allowed' worksheet.

    Layout (4 cols, NO merges):
      A1: title (cell A1 only, bold size 14)
      A2: column headers — Feed ID | Session | allowedPublisherIds | Notes (bold + light gray)
      A3+: per-feed groups:
           row: <feed_id> | (aggregate)  | sorted-union JSON or "(no data)" |
           row: <feed_id> | REGULAR      | JSON or "(no data)"              | optional "N passed + M top-up" note
           row: <feed_id> | PRE_MARKET   | …
           row: <feed_id> | POST_MARKET  | …
           row: <feed_id> | OVER_NIGHT   | …
           row: blank divider
      Footer: "Feeds skipped (no data for any mode):" then one feed_id per row in column A.
    """
    from openpyxl.styles import Font, PatternFill

    bold = Font(bold=True)
    bold_xl = Font(bold=True, size=14)
    gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    yellow = PatternFill(start_color="FFF4B5", end_color="FFF4B5", fill_type="solid")
    light_gray = PatternFill(
        start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
    )

    # Row 1: title (single cell, no merge).
    ws.cell(
        row=1, column=1, value=f"Allowed Publishers — {cluster} — {date}"
    ).font = bold_xl

    # Row 2: column headers.
    headers = ["Feed ID", "Session", "allowedPublisherIds", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = bold
        c.fill = gray

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:D2"  # extended to last data row at the end

    row = 3
    for feed_id, mode_data in per_feed_data.items():
        # Build per-session arrays (None if mode missing or no data after filter).
        per_session_arrays: list[list[int] | None] = []
        for mode in modes:
            md = mode_data.get(mode) if mode_data else None
            if md is None:
                per_session_arrays.append(None)
            else:
                # filtered is the selected list: threshold passers plus any below-threshold top-ups.
                ids = sorted({int(r["publisher_id"]) for r in md["filtered"]})
                per_session_arrays.append(ids if ids else None)

        # Aggregate row.
        agg = compute_aggregate(per_session_arrays)
        ws.cell(row=row, column=1, value=feed_id)
        ws.cell(row=row, column=2, value="(aggregate)")
        ws.cell(
            row=row,
            column=3,
            value=_format_allowed_pub_ids(agg) if agg else "(no data)",
        )
        if not agg:
            ws.cell(row=row, column=4, value="all sessions empty").fill = light_gray
        row += 1

        # Per-session rows.
        for mode, ids in zip(modes, per_session_arrays):
            session_label = sessions[mode]
            md = mode_data.get(mode) if mode_data else None
            ws.cell(row=row, column=1, value=feed_id)
            ws.cell(row=row, column=2, value=session_label)
            if md is None:
                ws.cell(row=row, column=3, value="(no data)")
                ws.cell(
                    row=row, column=4, value=f"mode missing for {date}"
                ).fill = light_gray
            elif ids is None:
                # Had data, but nothing passed and no publisher sat within the ceiling.
                ws.cell(row=row, column=3, value="(no data)")
                ws.cell(
                    row=row,
                    column=4,
                    value=f"0 passed, all > {_format_mult(ceiling_mult)}× ceiling",
                ).fill = light_gray
            else:
                ws.cell(row=row, column=3, value=_format_allowed_pub_ids(ids))
                if md["n_topup"] > 0:
                    ws.cell(
                        row=row,
                        column=4,
                        value=_topup_note(md["n_passed"], md["n_topup"], ceiling_mult),
                    ).fill = yellow
            row += 1

        row += 1  # blank divider between feeds

    # Skipped-feeds footer.
    if skipped_feeds:
        row += 1
        ws.cell(
            row=row, column=1, value="Feeds skipped (no data for any mode):"
        ).font = bold
        for fid in skipped_feeds:
            row += 1
            ws.cell(row=row, column=1, value=fid)

    # Update auto-filter range to include all data rows.
    last_data_row = max(row, 2)
    ws.auto_filter.ref = f"A2:D{last_data_row}"

    # Column widths.
    ws.column_dimensions["A"].width = 10  # Feed ID
    ws.column_dimensions["B"].width = 14  # Session
    ws.column_dimensions["C"].width = 80  # JSON snippet (full key + spaced array)
    ws.column_dimensions["D"].width = 32  # Notes


def _build_per_feed_data(
    feed_ids,
    reports_dir,
    cluster,
    date,
    excluded,
    top_n,
    max_ros_map,
    min_hit_map,
    min_obs,
    floor,
    ceiling_mult,
    modes,
):
    """Returns (per_feed_data, skipped_feeds, topup_rows, zero_passer_rows, modes_with_data_count).

    `modes` is the ordered list of dq_reports subdirectory names to read for each feed
    (drawn from ASSET_CLASS_CONFIG[<asset_class>]["modes"]).
    """
    per_feed_data: dict = {}
    skipped: list[int] = []
    topup_rows = 0
    zero_passer_rows = 0
    modes_with_data = 0

    for feed_id in feed_ids:
        mode_data: dict = {}
        any_data = False
        for mode in modes:
            raw = load_stats(reports_dir, cluster, mode, feed_id, date)
            if raw is None:
                mode_data[mode] = None
                continue
            # Apply exclusion at the row level.
            kept = []
            for r in raw:
                try:
                    pid = int(r["publisher_id"])
                except (ValueError, KeyError):
                    continue
                if pid in excluded:
                    continue
                kept.append(r)
            if not kept:
                mode_data[mode] = None  # all rows excluded
                continue
            ranked = rank_top_n(kept, n=top_n, excluded=set())  # already excluded
            selected, n_passed, n_topup = apply_filter(
                kept, max_ros_map[mode], min_hit_map[mode], min_obs, floor, ceiling_mult
            )
            mode_data[mode] = {
                "ranked": ranked,
                "filtered": selected,
                "n_passed": n_passed,
                "n_topup": n_topup,
            }
            any_data = True
            modes_with_data += 1
            if n_topup > 0:
                topup_rows += 1
            if n_passed == 0:
                zero_passer_rows += 1
        if not any_data:
            skipped.append(feed_id)
        per_feed_data[feed_id] = mode_data
    return per_feed_data, skipped, topup_rows, zero_passer_rows, modes_with_data


def main():
    parser = argparse.ArgumentParser(
        description="Generate one Excel summary workbook from evaluate_feeds_bulk DQ outputs.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  # Default (us-equities, 4 modes):
  python3 -m lazer_dq.summarize_feeds \\
      --csv MV_Mario_3_pre.csv --cluster lazer-prod --date 2026-05-06

  # HK equities:
  python3 -m lazer_dq.summarize_feeds \\
      --csv equity_hk_feed_ids.csv --asset-class hk-equities \\
      --cluster lazer-prod --date 2026-05-19
""",
    )
    parser.add_argument(
        "--csv", required=True, help="CSV: feed_id,date,mode per row (column 1 used)"
    )
    parser.add_argument(
        "--cluster", required=True, help="Cluster name (e.g. lazer-prod)"
    )
    parser.add_argument("--date", required=True, help="Date YYYY-MM-DD")
    parser.add_argument(
        "--reports-dir",
        default="dq_reports",
        help="Base reports directory (default: dq_reports)",
    )
    parser.add_argument(
        "--publishers-md",
        default="publishers.md",
        help="Path to publishers.md (default: publishers.md)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path (default: dq_summary_<cluster>_<date>.xlsx)",
    )
    parser.add_argument(
        "--asset-class",
        choices=sorted(ASSET_CLASS_CONFIG.keys()),
        default="us-equities",
        help="Asset class to summarize (default: us-equities). Determines which "
        "dq_reports/<cluster>/<mode>/ directories are read and the workbook layout.",
    )
    parser.add_argument(
        "--max-rmse-over-spread-regular",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_max_ros"]["us-equities"],
    )
    parser.add_argument(
        "--min-hit-rate-regular",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_min_hit"]["us-equities"],
    )
    parser.add_argument(
        "--max-rmse-over-spread-pre",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_max_ros"]["us-equities-pre"],
    )
    parser.add_argument(
        "--min-hit-rate-pre",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_min_hit"]["us-equities-pre"],
    )
    parser.add_argument(
        "--max-rmse-over-spread-post",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_max_ros"][
            "us-equities-post"
        ],
    )
    parser.add_argument(
        "--min-hit-rate-post",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_min_hit"][
            "us-equities-post"
        ],
    )
    parser.add_argument(
        "--max-rmse-over-spread-overnight",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_max_ros"][
            "us-equities-overnight"
        ],
    )
    parser.add_argument(
        "--min-hit-rate-overnight",
        type=float,
        default=ASSET_CLASS_CONFIG["us-equities"]["default_min_hit"][
            "us-equities-overnight"
        ],
    )
    parser.add_argument("--min-n-observations", type=int, default=DEFAULT_MIN_N_OBS)
    parser.add_argument("--top-n", type=int, default=DEFAULT_TOP_N)
    parser.add_argument(
        "--redundancy-floor",
        type=int,
        default=DEFAULT_REDUNDANCY_FLOOR,
        help="Minimum publishers per feed/session; top up below-threshold "
        "near-misses to reach it (default: 5).",
    )
    parser.add_argument(
        "--topup-ceiling-mult",
        type=float,
        default=DEFAULT_TOPUP_CEILING_MULT,
        help="A top-up's rmse_over_spread must be <= this multiple of the "
        "per-mode pass threshold (default: 2.0).",
    )
    args = parser.parse_args()

    csv_path = Path(args.csv)
    md_path = Path(args.publishers_md)
    reports_dir = Path(args.reports_dir)

    if not csv_path.exists():
        print(f"Error: CSV file '{csv_path}' not found.")
        sys.exit(1)
    if not md_path.exists():
        print(
            f"Error: publishers.md '{md_path}' not found (needed for .Test exclusion)."
        )
        sys.exit(1)
    if not (reports_dir / args.cluster).exists():
        print(f"Error: reports dir '{reports_dir / args.cluster}' not found.")
        sys.exit(1)

    excluded = load_excluded_publishers(md_path)

    asset_cfg = ASSET_CLASS_CONFIG[args.asset_class]
    modes = asset_cfg["modes"]
    sessions = asset_cfg["sessions"]

    validate_csv_modes(csv_path, allowed_modes=modes)

    feed_ids = discover_feeds(csv_path)
    if not feed_ids:
        print(f"Error: no feed_ids parsed from '{csv_path}'.")
        sys.exit(1)

    if args.asset_class == "us-equities":
        # us-equities keeps its existing flat per-mode CLI flags.
        max_ros_map = {
            "us-equities": args.max_rmse_over_spread_regular,
            "us-equities-pre": args.max_rmse_over_spread_pre,
            "us-equities-post": args.max_rmse_over_spread_post,
            "us-equities-overnight": args.max_rmse_over_spread_overnight,
        }
        min_hit_map = {
            "us-equities": args.min_hit_rate_regular,
            "us-equities-pre": args.min_hit_rate_pre,
            "us-equities-post": args.min_hit_rate_post,
            "us-equities-overnight": args.min_hit_rate_overnight,
        }
    else:
        # Other asset classes use the registry defaults (no per-mode CLI overrides yet).
        max_ros_map = dict(asset_cfg["default_max_ros"])
        min_hit_map = dict(asset_cfg["default_min_hit"])

    (
        per_feed_data,
        skipped,
        topup_rows,
        zero_passer_rows,
        modes_with_data,
    ) = _build_per_feed_data(
        feed_ids,
        reports_dir,
        args.cluster,
        args.date,
        excluded,
        args.top_n,
        max_ros_map,
        min_hit_map,
        args.min_n_observations,
        args.redundancy_floor,
        args.topup_ceiling_mult,
        modes=modes,
    )

    feeds_with_data = len(feed_ids) - len(skipped)
    if feeds_with_data == 0:
        print("Error: no feed produced any data (wrong --date or --cluster?).")
        sys.exit(1)

    # Build workbook.
    from openpyxl import Workbook

    wb = Workbook()
    ws_rank = wb.active
    ws_rank.title = "rankings"
    ws_allow = wb.create_sheet("allowed")
    write_rankings_sheet(ws_rank, per_feed_data, args.date, args.cluster, modes=modes)
    write_allowed_sheet(
        ws_allow,
        per_feed_data,
        skipped,
        args.date,
        args.cluster,
        modes=modes,
        sessions=sessions,
        ceiling_mult=args.topup_ceiling_mult,
    )

    out_path = (
        Path(args.output)
        if args.output
        else Path(f"dq_summary_{args.cluster}_{args.date}.xlsx")
    )
    wb.save(out_path)

    test_count = sum(1 for _ in excluded if _ != 0)
    sample_excluded = sorted(p for p in excluded if p != 0)[:3]
    print(f"Summary written to {out_path}")
    print(f"Feeds in CSV: {len(feed_ids)}")
    print(f"Feeds with at least one mode: {feeds_with_data}")
    if skipped:
        print(f"Feeds skipped (no data anywhere): {len(skipped)} → {skipped}")
    else:
        print("Feeds skipped (no data anywhere): 0")
    print(f"Modes with data: {modes_with_data}/{len(feed_ids) * len(modes)} cells")
    print(f"Excluded publishers: 0 + {test_count} .Test (sample: {sample_excluded})")
    print(f"Rows using top-ups: {topup_rows} cells")
    print(f"Rows with 0 passers: {zero_passer_rows} cells")
    sys.exit(0)


if __name__ == "__main__":
    main()
